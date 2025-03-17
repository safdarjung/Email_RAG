#!/usr/bin/env python3
from __future__ import print_function
import os
import re
import json
import pickle
import base64
import csv
import time
import logging
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from typing import Any, Dict, List, Optional, Tuple, ClassVar
import contextlib
import io
import functools # Import for functools.partial
import shutil # For file copying

from dotenv import load_dotenv
from pydantic import Field
from crewai import LLM, Agent
from crewai.tools.base_tool import BaseTool

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

from PyPDF2 import PdfReader
import openpyxl
import docx

from litellm.exceptions import RateLimitError

try:
    from PIL import Image, ImageTk
    import pytesseract
except ImportError:
    pytesseract = None

try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib

matplotlib.use('TkAgg') # Ensure TkAgg backend is used for matplotlib with Tkinter

# For fuzzy matching of file names
try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None

# For in-memory embedding-based retrieval
try:
    from sentence_transformers import SentenceTransformer, util
    EMBEDDING_MODEL = SentenceTransformer('all-MiniLM-L6-v2')
except ImportError:
    EMBEDDING_MODEL = None

# Load environment variables and set scopes
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

###############################################################################
# Global LLM instance (unified_llm)
###############################################################################
unified_llm = LLM(
    model="gemini/gemini-2.0-flash",  # Or "gemini-pro" if flash is unavailable
    api_key=GEMINI_API_KEY,
    temperature=0.1
)

###############################################################################
# Function: llm_predict_with_backoff
###############################################################################
def llm_predict_with_backoff(prompt: str, max_retries: int = 5, initial_delay: float = 1.0) -> str:
    """Calls the unified LLM with exponential backoff."""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            time.sleep(2.0)
            response = unified_llm.call(prompt)
            return response
        except RateLimitError as e:
            logging.warning("Rate limit error on attempt %d/%d: %s", attempt+1, max_retries, e)
            time.sleep(delay)
            delay *= 2
        except Exception as e:
            logging.error("Error during LLM call: %s", e)
            return ""
    raise RateLimitError("Exceeded maximum retry attempts for LLM call")

###############################################################################
# Function: extract_text_from_file
###############################################################################
def extract_text_from_file(file_path: str) -> str:
    """
    Extracts text from a file (PDF, Excel, Word, CSV, image, or plain text).
    """
    text = ""
    fname = file_path.lower()
    try:
        if fname.endswith('.pdf'):
            with open(file_path, 'rb') as file_content:
                reader = PdfReader(file_content)
                extracted = ""
                for page in reader.pages:
                    pg_text = page.extract_text()
                    if pg_text:
                        extracted += pg_text + "\n"
            if len(extracted.strip()) < 50 and convert_from_path:
                pages = convert_from_path(file_path, dpi=200)
                if pytesseract:
                    for p in pages:
                        extracted += pytesseract.image_to_string(p) + "\n"
            text = extracted
        elif fname.endswith(('.xlsx', '.xls')):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append("\t".join(str(cell) if cell is not None else "" for cell in row))
            text = "\n".join(rows)
        elif fname.endswith('.docx'):
            document = docx.Document(file_path)
            text = "\n".join(p.text for p in document.paragraphs)
        elif fname.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8') as file_content:
                reader = csv.reader(file_content)
                rows = []
                for row in reader:
                    rows.append("\t".join(row))
                text = "\n".join(rows)
        elif fname.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
            if pytesseract:
                image = Image.open(file_path)
                text = pytesseract.image_to_string(image)
            else:
                text = "[No OCR support installed]"
        else:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
        return text
    except Exception as e:
        logging.error("Error extracting text from %s: %s", file_path, e)
        return f"Error extracting text from {file_path}: {e}"

###############################################################################
# Function: interpret_user_query
###############################################################################
def interpret_user_query(query: str) -> Dict[str, Any]:
    """
    Uses the LLM to convert a natural language query into a Gmail search query.
    Returns a dict with keys: gmail_query, fetch_mode, and has_attachments.
    """
    prompt = (
        f"Interpret the following user query and convert it into a valid Gmail search query by understanding the user's intent. "
        f"Use one of the following formats as appropriate:\n"
        f"   - from:example@gmail.com has:attachment\n"
        f"   - to:recipient@gmail.com after:2023/01/01 before:2023/12/31\n"
        f"   - subject:(\"meeting notes\" OR \"project update\")\n"
        f"   - is:unread label:work\n"
        f"   - \"important information\" -subject:\"important information\"\n"
        f"   - is:starred subject:report\n"
        f"   - in:spam from:spammer@example.com\n"
        f"   - has:attachment filename:pdf\n"
        f"   - \"follow up\" before:2023/01/01\n\n"
        f"Also, decide if the fetch_mode should be 'summary', 'full', or 'analyze', and set has_attachments to true if attachments are expected.\n\n"
        f"User Query: \"{query}\"\n"
        f"Return only the JSON with keys 'gmail_query', 'fetch_mode', and 'has_attachments'."
    )
    result = llm_predict_with_backoff(prompt)
    try:
        result = result.replace("`json", "").replace("`", "").strip()
        data = json.loads(result)
        if not any(op in data.get("gmail_query", "").lower() for op in ["from:", "to:", "after:", "before:", "subject:", "is:", "label:", "has:"]) and "\"" not in data.get("gmail_query", ""):
            if data.get("gmail_query", "").strip():
                data["gmail_query"] = f"\"{data['gmail_query']}\""
            else:
                raise ValueError("Interpreted query does not appear valid.")
        return data
    except Exception as e:
        logging.error("Error interpreting query: %s", e)
        return {
            "gmail_query": f"\"{query}\"",
            "fetch_mode": "summary",
            "has_attachments": "attachment" in query.lower()
        }

###############################################################################
# Function: fetch_gmail_data
###############################################################################
def fetch_gmail_data(query: Optional[str] = None, batch_size: int = 50) -> List[Dict[str, Any]]:
    """
    Uses Gmail API (with OAuth) to fetch email messages matching the query.
    Returns a list of message dictionaries.
    """
    creds = None
    token_path = 'token.pickle'
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'wb') as token:
            pickle.dump(creds, token)
    service = build('gmail', 'v1', credentials=creds)
    fetched_messages = []
    page_token = None
    while len(fetched_messages) < batch_size:
        if query:
            results = service.users().messages().list(
                userId='me',
                q=query,
                maxResults=min(batch_size, 100),
                pageToken=page_token
            ).execute()
        else:
            results = service.users().messages().list(
                userId='me',
                labelIds=['INBOX'],
                maxResults=min(batch_size, 100),
                pageToken=page_token
            ).execute()
        messages = results.get('messages', [])
        page_token = results.get('nextPageToken')
        for message in messages:
            msg = service.users().messages().get(
                userId='me',
                id=message['id'],
                format='full'
            ).execute()
            fetched_messages.append(msg)
            if len(fetched_messages) >= batch_size:
                break
        if not page_token:
            break
        time.sleep(1)
    return fetched_messages

###############################################################################
# Function: edit_file_content
###############################################################################
def edit_file_content(initial_content: str, history: str = "") -> str:
    """
    Opens a dialog for editing or analyzing file content.
    If conversation history is provided, it is included in the prompt.
    Returns the edited/analyzed content.
    """
    edit_win = tk.Toplevel()
    edit_win.title("Edit File Content / Analyze Data")
    edit_win.geometry("600x400")

    tk.Label(edit_win, text="Current Content (read-only):").pack(anchor="w", padx=10, pady=(10, 0))
    content_text = ScrolledText(edit_win, wrap=tk.WORD, height=10)
    content_text.pack(fill="both", expand=True, padx=10, pady=5)
    content_text.insert("1.0", initial_content)
    content_text.configure(state=tk.DISABLED)

    tk.Label(edit_win, text="Enter your instruction:").pack(anchor="w", padx=10, pady=(10, 0))
    instruction_entry = tk.Text(edit_win, wrap=tk.WORD, height=4)
    instruction_entry.pack(fill="x", padx=10, pady=5)

    result = {"edited": initial_content}

    def submit_edit():
        instruction = instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            messagebox.showerror("Error", "Instruction cannot be empty.")
            return
        prompt_context = f"Conversation history:\n{history}\n\n" if history else ""
        prompt = (
            prompt_context +
            "You are an advanced text editor, data analyst, and visualization assistant. "
            "If the instruction requires analysis or visualization, output the charts/graphs directly. "
            "Do not return raw Python code for visualization; instead, display the actual chart.\n\n"
            f"Instruction: {instruction}\n\n"
            f"Content:\n{initial_content}\n\n"
            "Return the final text, analysis, and if applicable, display the chart."
        )
        try:
            response = unified_llm.call(prompt)
            final_response = process_visualization_response(response)
            if not final_response.strip():
                messagebox.showinfo("LLM Output", "LLM returned no output.")
            result["edited"] = final_response.strip()
            edit_win.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to edit/analyze content: {e}")

    submit_btn = ttk.Button(edit_win, text="Submit Instruction", command=submit_edit)
    submit_btn.pack(padx=10, pady=10)

    edit_win.wait_window()
    return result["edited"]

###############################################################################
# Function: custom_chunk_text
###############################################################################
def custom_chunk_text(text: str, method: str) -> List[str]:
    """Splits text into chunks based on the specified method."""
    if method == "word":
        words = text.split()
        return [" ".join(words[i:i+50]) for i in range(0, len(words), 50)]
    elif method == "sentence":
        sentences = re.split(r'(?<=[.!?])\s+', text)
        return sentences
    elif method == "paragraph":
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        return paragraphs
    elif method == "pages":
        pages = re.split(r'\n{2,}', text)
        return pages
    elif method == "batches":
        pages = re.split(r'\n{2,}', text)
        batch_size = 3
        return ["\n\n".join(pages[i:i+batch_size]) for i in range(0, len(pages), batch_size)]
    elif method == "headings":
        lines = text.splitlines()
        chunks = []
        current_chunk = []
        for line in lines:
            if line.isupper() and len(line.strip()) > 3:
                if current_chunk:
                    chunks.append("\n".join(current_chunk))
                    current_chunk = []
                current_chunk.append(line)
            else:
                current_chunk.append(line)
        if current_chunk:
            chunks.append("\n".join(current_chunk))
        return chunks
    elif method == "headings_subheadings":
        lines = text.splitlines()
        chunks = []
        current_chunk = []
        for line in lines:
            if (line.isupper() and len(line.strip()) > 3) or (line.istitle() and len(line.split()) < 5):
                if current_chunk:
                    chunks.append("\n".join(current_chunk))
                    current_chunk = []
                current_chunk.append(line)
            else:
                current_chunk.append(line)
        if current_chunk:
            chunks.append("\n".join(current_chunk))
        return chunks
    else:
        return [text]


###############################################################################
# ScrollableFrame for Gmail Emails Panel
###############################################################################
class ScrollableFrame(ttk.Frame):
    """
    A reusable scrollable frame. Place widgets into self.scrollable_frame.
    """
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

###############################################################################
# DocumentStore with File Renaming and Mapping
###############################################################################
class DocumentStore:
    def __init__(self, use_embeddings: bool = True, chunk_size: int = 500):
        self.use_embeddings = use_embeddings and (EMBEDDING_MODEL is not None)
        self.chunk_size = chunk_size
        self.docs: List[Tuple[str, str, str]] = []  # (short_name, chunk, original_name) - added original_name to each chunk
        self.embeddings = []  # Embeddings for each chunk
        self.short_to_original: Dict[str, str] = {}
        self.original_to_short: Dict[str, str] = {}
        self.file_counter = 1

    def _generate_short_name(self, original_name: str) -> str:
        short_name = f"File_{self.file_counter}"
        self.file_counter += 1
        self.short_to_original[short_name] = original_name
        self.original_to_short[original_name] = short_name
        return short_name

    def add_document_text(self, text: str, original_name: str):
        short_name = self._generate_short_name(original_name)
        chunks = self._chunk_text(text, self.chunk_size)
        for chunk in chunks:
            self.docs.append((short_name, chunk, original_name)) # Store original_name with chunk
            if self.use_embeddings:
                emb = EMBEDDING_MODEL.encode(chunk, convert_to_tensor=True)
                self.embeddings.append(emb)

    def _chunk_text(self, text: str, chunk_size: int) -> List[str]:
        chunks = []
        start = 0
        while start < len(text):
            end = min(start + chunk_size, len(text))
            chunks.append(text[start:end])
            start = end
        return chunks

    def retrieve_relevant_chunks(self, query: str, top_k: int = 3) -> List[Tuple[str, str, str]]: # Return tuple (short_name, chunk, original_name)
        # Fuzzy matching to find a referenced file
        if fuzz is not None:
            best_score = 0
            best_original = None
            for orig in self.original_to_short:
                score = fuzz.partial_ratio(orig.lower(), query.lower())
                if score > best_score:
                    best_score = score
                    best_original = orig
            if best_score > 70:
                short_name = self.original_to_short[best_original]
                return self._retrieve_from_short_name(short_name, query, top_k)
        else:
            for orig in self.original_to_short:
                if orig.lower() in query.lower():
                    short_name = self.original_to_short[orig]
                    return self._retrieve_from_short_name(short_name, query, top_k)
        # Fallback: retrieve from all docs
        if self.use_embeddings:
            query_emb = EMBEDDING_MODEL.encode(query, convert_to_tensor=True)
            scores = util.pytorch_cos_sim(query_emb, self.embeddings)[0]
            top_results = scores.topk(k=min(top_k, len(scores))).indices.tolist()
            return [self.docs[idx] for idx in top_results] # Return tuples now
        else:
            tokens_query = set(query.lower().split())
            def score(item: Tuple[str, str, str]) -> int: # Score now works with tuple
                return len(tokens_query.intersection(set(item[1].lower().split())))
            ranked = sorted(self.docs, key=score, reverse=True)
            return [doc for doc in ranked[:top_k]] # Return tuples

    def _retrieve_from_short_name(self, short_name: str, query: str, top_k: int) -> List[Tuple[str, str, str]]: # Return tuples
        relevant_docs = [(sn, chunk, oname) for (sn, chunk, oname) in self.docs if sn == short_name] # Include original_name
        if not relevant_docs:
            return [("","", f"No chunks found for file {short_name}")] # Return tuple with empty short_name, chunk, and error in original_name
        if self.use_embeddings:
            relevant_indices = [i for i, (sn, _, _) in enumerate(self.docs) if sn == short_name]
            relevant_chunks = [self.docs[i][1] for i in relevant_indices]
            relevant_original_names = [self.docs[i][2] for i in relevant_indices] # Get original names too
            relevant_embs = [self.embeddings[i] for i in relevant_indices]
            query_emb = EMBEDDING_MODEL.encode(query, convert_to_tensor=True)
            scores = util.pytorch_cos_sim(query_emb, relevant_embs)[0]
            top_results = scores.topk(k=min(top_k, len(scores))).indices.tolist()
            return [(short_name, relevant_chunks[idx], relevant_original_names[idx]) for idx in top_results] # Return tuples with original_name
        else:
            tokens_query = set(query.lower().split())
            def score(item: Tuple[str, str, str]) -> int: # Score works with tuple
                return len(tokens_query.intersection(set(item[1].lower().split())))
            ranked = sorted(relevant_docs, key=score, reverse=True)
            return [doc for doc in ranked[:top_k]] # Return tuples


###############################################################################
# RAGChatBot
###############################################################################
class RAGChatBot:
    def __init__(self, doc_store: DocumentStore):
        self.doc_store = doc_store
        self.conversation_history = ""
        self.retrieved_context_content = "" # Store last retrieved context

    def add_document(self, text: str, original_name: str):
        self.doc_store.add_document_text(text, original_name)

    def chat(self, user_query: str, top_k: int = 3) -> str:
        relevant_chunks_with_names = self.doc_store.retrieve_relevant_chunks(user_query, top_k=top_k) # Now get tuples
        context_block = "\n\n".join([f"Relevant chunk from '{original_name}':\n{chunk}" for _, chunk, original_name in relevant_chunks_with_names]) # Use original_name from tuple
        self.retrieved_context_content = context_block # Store for display
        prompt = (
            "You are a helpful assistant with access to local documents via the doc store. "
            "If the user references a file by name, use that file's data.\n\n"
            f"{context_block}\n\n"
            f"User: {user_query}\nAssistant:"
        )
        response = unified_llm.call(prompt)
        self.conversation_history += f"\nUser: {user_query}\nAssistant: {response}"
        return process_visualization_response(response)

    def get_last_retrieved_context(self) -> str:
        return self.retrieved_context_content

###############################################################################
# Function: verify_and_run_code (for visualization code) - UPDATED
###############################################################################
def verify_and_run_code(code: str, namespace: Dict[str, Any]) -> str:
    """Executes code in a provided namespace and displays a matplotlib figure if generated.
       Captures stdout and stderr and returns them as part of the output.
    """
    output_buffer = io.StringIO() # Capture stdout
    error_buffer = io.StringIO()  # Capture stderr
    namespace['plt'] = plt # Ensure plt is in namespace

    try:
        with contextlib.redirect_stdout(output_buffer), contextlib.redirect_stderr(error_buffer): # Redirect both
            exec(code, namespace)
        fig = plt.gcf() # Check for figure *after* execution in case code creates it.
        chart_win = tk.Toplevel()
        chart_win.title("Generated Visualization")
        canvas = FigureCanvasTkAgg(fig, master=chart_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        plt.close(fig) # Close figure to prevent display in main window or memory leak.
        output_text = output_buffer.getvalue() # Get stdout
        error_text = error_buffer.getvalue()   # Get stderr
        full_output = "[Visualization Generated]\n" + output_text + ("\n[Errors]:\n" + error_text if error_text else "") # Include errors if any

        return full_output

    except Exception as e:
        error_text = error_buffer.getvalue() # Capture any stderr even on exception
        error_message = f"[Error executing code: {e}]\n[Standard Error output]:\n{error_text}" if error_text else f"[Error executing code: {e}]"

        logging.error(f"Error in verify_and_run_code: {error_message}", exc_info=True) # Log full exception info

        return error_message

###############################################################################
# Function: process_visualization_response
###############################################################################
def process_visualization_response(response: str) -> str:
    """
    Checks if the response contains a code block delimited by triple backticks.
    If yes, indicates code block found for visualization.
    Also, if "CHART:" is found, generates a dummy chart.
    """
    code_blocks = re.findall(r"`(.*?)`", response, re.DOTALL)
    if code_blocks:
        return "[Code for visualization provided in code editor]" # Indicate code block found.
    if "CHART:" in response:
        chart_type = "BAR"  # Dummy; update as needed.
        x = [1, 2, 3, 4, 5]
        y = [10, 20, 15, 30, 25]
        fig, ax = plt.subplots()
        if chart_type.upper() == "BAR":
            ax.bar(x, y, color='skyblue')
        elif chart_type.upper() == "LINE":
            ax.plot(x, y, marker='o')
        else:
            ax.plot(x, y, marker='o')
        ax.set_title("Generated Chart")
        ax.set_xlabel("X-axis")
        ax.set_ylabel("Y-axis")
        chart_win = tk.Toplevel()
        chart_win.title("Chart")
        canvas = FigureCanvasTkAgg(fig, master=chart_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        return response.replace("CHART:", "[Chart generated above]")
    return response

###############################################################################
# Custom Editors
###############################################################################
class CustomTextEditor(tk.Toplevel):
    def __init__(self, parent, file_path, master_app): # Added master_app
        super().__init__(parent)
        self.master_app = master_app # Store master app reference
        self.title(f"Text Editor - {os.path.basename(file_path)}")
        self.geometry("1000x800") # Increased size

        self.file_path = file_path
        self.current_content = extract_text_from_file(file_path) # Store content

        # --- Paned Window for Split View ---
        self.paned_window = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        # --- Editor and NL Instruction Frame (LEFT) ---
        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2) # Wider for editor

        # ---- Text Area ----
        self.text_area = ScrolledText(left_frame, wrap=tk.WORD, height=25)
        self.text_area.pack(expand=True, fill='both', padx=10, pady=(10,0))
        self.text_area.insert('1.0', self.current_content)

        # ---- NL Instruction Frame ----
        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)
        self.save_button = ttk.Button(nl_frame, text="Save Changes", command=self.save_changes_to_file) # SAVE Button
        self.save_button.pack(pady=5)


        # --- RAG Agent Output and Context Frame (RIGHT) ---
        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        # ---- RAG Agent Output Window ----
        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED) # Read-only
        self.rag_output_display.pack(expand=True, fill='both')

        # ---- Context Access Window ----
        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent") # Context Frame
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED) # Read-only, smaller height
        self.context_display.pack(expand=True, fill='both')


        self.rag_agent = RAGTextEditorAgent(self) # Initialize RAG agent for this editor

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return

        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("") # Clear context display on new instruction

        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            response = self.rag_agent.run_instruction(instruction, self.current_content) # Pass current content
            self.set_rag_output(response)

            # Update text area if the agent returns modified content (you can adjust logic here)
            if "[EDITED_CONTENT]" in response:
                edited_content = response.split("[EDITED_CONTENT]")[1].strip() # Simple split, improve parsing if needed
                self.update_text_content(edited_content) # Method to update text area and current_content

            context_content = self.master_app.chatbot.get_last_retrieved_context() # Get context from chatbot
            self.set_context_output(context_content) # Display context

        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")


    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL) # Enable editing to insert
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED) # Disable editing

    def set_context_output(self, text): # New method for context display
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)

    def update_text_content(self, new_content):
        self.text_area.delete("1.0", tk.END)
        self.text_area.insert("1.0", new_content)
        self.current_content = new_content # Update stored content

    def save_changes_to_file(self): # SAVE FUNCTIONALITY
        new_content = self.text_area.get("1.0", tk.END)
        try:
            with open(self.file_path, 'w', encoding='utf-8') as f: # Save to original file path
                f.write(new_content)
            self.current_content = new_content # Update stored content
            messagebox.showinfo("File Saved", f"Changes saved to: {os.path.basename(self.file_path)}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Error saving file: {e}")


class CustomImageViewer(tk.Toplevel):
    def __init__(self, parent, file_path, master_app): # Added master_app
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"Image Viewer - {os.path.basename(file_path)}")
        self.geometry("900x700")

        self.file_path = file_path

        # --- Paned Window for Split View ---
        self.paned_window = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        # --- Image Display and NL Frame (LEFT) ---
        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        # ---- Image Label ----
        try:
            img = Image.open(file_path)
            img_tk = ImageTk.PhotoImage(img)
            self.image_label = tk.Label(left_frame, image=img_tk)
            self.image_label.image = img_tk # Keep a reference
            self.image_label.pack(padx=10, pady=(10,0))
        except Exception as e:
            tk.Label(left_frame, text=f"Error opening image: {e}").pack(padx=10, pady=10)

        # ---- NL Instruction Frame ----
        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)

        # --- RAG Agent Output and Context Frame (RIGHT) ---
        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        # ---- RAG Agent Output Window ----
        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED) # Read-only
        self.rag_output_display.pack(expand=True, fill='both')

        # ---- Context Access Window ----
        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent") # Context Frame
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED) # Read-only, smaller height
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGFileViewerAgent(self) # Generic RAG agent for file viewers

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return

        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("") # Clear context

        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            file_content = extract_text_from_file(self.file_path) # Extract fresh content for analysis
            response = self.rag_agent.run_instruction(instruction, file_content)
            self.set_rag_output(response)

            context_content = self.master_app.chatbot.get_last_retrieved_context() # Get context from chatbot
            self.set_context_output(context_content) # Display context

        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")


    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL) # Enable editing to insert
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED) # Disable editing

    def set_context_output(self, text): # New method for context display
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)


class CustomExcelViewer(tk.Toplevel):
    def __init__(self, parent, file_path, master_app): # Added master_app
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"Excel Viewer - {os.path.basename(file_path)}")
        self.geometry("900x700")

        self.file_path = file_path

        # --- Paned Window for Split View ---
        self.paned_window = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        # --- Excel Treeview and NL Frame (LEFT) ---
        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        # ---- Excel Treeview ----
        self.tree = ttk.Treeview(left_frame)
        self.tree.pack(expand=True, fill='both', padx=10, pady=(10,0))
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            header = [cell.value for cell in sheet[1]]
            self.tree["columns"] = header
            self.tree["show"] = "headings"
            for col in header:
                self.tree.heading(col, text=col)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", tk.END, values=row)
        except Exception as e:
            tk.Label(left_frame, text=f"Error opening Excel file: {e}").pack(padx=10, pady=10)

        # ---- NL Instruction Frame ----
        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)

        # --- RAG Agent Output and Context Frame (RIGHT) ---
        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        # ---- RAG Agent Output Window ----
        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED) # Read-only
        self.rag_output_display.pack(expand=True, fill='both')

        # ---- Context Access Window ----
        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent") # Context Frame
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED) # Read-only, smaller height
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGFileViewerAgent(self) # Generic RAG agent

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return

        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("") # Clear context

        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            file_content = extract_text_from_file(self.file_path) # Extract fresh content
            response = self.rag_agent.run_instruction(instruction, file_content)
            self.set_rag_output(response)

            context_content = self.master_app.chatbot.get_last_retrieved_context() # Get context from chatbot
            self.set_context_output(context_content) # Display context

        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")


    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL) # Enable editing to insert
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED) # Disable editing

    def set_context_output(self, text): # New method for context display
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)


class CustomPDFViewer(tk.Toplevel):
    def __init__(self, parent, file_path, master_app): # Added master_app
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"PDF Viewer - {os.path.basename(file_path)}")
        self.geometry("900x700")

        self.file_path = file_path
        self.image_based_pdf = False # Placeholder for future image-based PDF handling

        # --- Paned Window for Split View ---
        self.paned_window = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        # --- PDF Text Area and NL Frame (LEFT) ---
        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        # ---- PDF Text Area ----
        self.text_area = ScrolledText(left_frame, wrap=tk.WORD, height=25)
        self.text_area.pack(expand=True, fill='both', padx=10, pady=(10,0))
        content = extract_text_from_file(file_path)
        self.text_area.insert('1.0', content)
        self.text_area.config(state=tk.DISABLED) # Read-only for PDF viewer

        # ---- NL Instruction Frame ----
        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)

        # --- RAG Agent Output and Context Frame (RIGHT) ---
        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        # ---- RAG Agent Output Window ----
        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED) # Read-only
        self.rag_output_display.pack(expand=True, fill='both')

        # ---- Context Access Window ----
        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent") # Context Frame
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED) # Read-only, smaller height
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGFileViewerAgent(self) # Generic RAG agent

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return

        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("") # Clear context

        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            file_content = extract_text_from_file(self.file_path) # Extract fresh content
            response = self.rag_agent.run_instruction(instruction, file_content)
            self.set_rag_output(response)

            context_content = self.master_app.chatbot.get_last_retrieved_context() # Get context from chatbot
            self.set_context_output(context_content) # Display context

        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")


    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL) # Enable editing to insert
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED) # Disable editing

    def set_context_output(self, text): # New method for context display
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)


def open_custom_editor(parent, file_path, master_app): # Added master_app
    if file_path.lower().endswith(('.txt', '.csv', '.docx')): # Added docx as text-editable
        CustomTextEditor(parent, file_path, master_app) # Pass master_app
    elif file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
        CustomImageViewer(parent, file_path, master_app) # Pass master_app
    elif file_path.lower().endswith(('.xlsx', '.xls')):
        CustomExcelViewer(parent, file_path, master_app) # Pass master_app
    elif file_path.lower().endswith('.pdf'):
        CustomPDFViewer(parent, file_path, master_app) # Pass master_app
    else:
        messagebox.showinfo("Info", "No custom viewer available for this file type. Opening as text editor.")
        CustomTextEditor(parent, file_path, master_app) # Pass master_app


###############################################################################
# RAG Agents - UPDATED - Added RAGFileViewerAgent and RAGDropboxAgent
###############################################################################
class RAGTextEditorAgent:
    def __init__(self, editor: CustomTextEditor):
        self.editor = editor # Reference to the editor window

    def run_instruction(self, instruction: str, file_content: str) -> str:
        prompt = (
            "You are an expert text editor AI. Analyze the content and follow the user's instruction.\n"
            "You have access to relevant context from emails and dropbox files if applicable. Use this context to better understand the user's request and file content.\n"
            "If the instruction is to edit the content, perform the edit and clearly indicate the edited content in your response like this: '[EDITED_CONTENT]\n[the edited content here]\n[/EDITED_CONTENT]'. "
            "If the instruction is for analysis, perform the analysis and present the result clearly.\n"
            f"Instruction: {instruction}\n"
            f"File Content:\n{file_content}\n"
            "Response:"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw) # Process for charts if needed
        return response_processed


class RAGFileViewerAgent: # Generic agent for viewers (Image, Excel, PDF)
    def __init__(self, viewer): # Can be any viewer type
        self.viewer = viewer

    def run_instruction(self, instruction: str, file_content: str) -> str:
        prompt = (
            "You are an expert AI assistant for file analysis. You are viewing a file and will answer user's questions or perform analysis based on the file content.\n"
            "You have access to relevant context from emails and dropbox files if applicable. Use this context to better understand the user's request and file content.\n"
            "Analyze the file content and provide a helpful and informative response to the user's instruction.\n"
            "If the file is an image, perform OCR if possible and analyze the text. If it's a spreadsheet, analyze the data. If it's a PDF, analyze the text content.\n"
            f"Instruction: {instruction}\n"
            f"File Content:\n{file_content}\n"
            "Response:"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw) # Process for charts
        return response_processed


class RAGCodeEditorAgent:
    def __init__(self, editor: 'AICodeEditor'): # Forward reference, needs string for class definition order
        self.editor = editor

    def run_instruction(self, instruction: str, code_content: str, file_context: str = "") -> str:
        context_prompt_part = f"Context from your files:\n{file_context}\n\n" if file_context else ""
        prompt = (
            "You are an expert AI code editor and data visualization assistant. "
            "You can generate, analyze, debug, and run Python code. "
            "You can also create data visualizations using matplotlib.pyplot. \n"
            "You have access to the content of files loaded in the system (see context below if available), emails, and dropbox files. "
            "Use any relevant information from this context to fulfill the user's instructions. Do not access external websites or local file system other than provided file context.\n"
            "If the user asks for code, provide valid, runnable Python code. If asked to visualize data, generate Python code to create charts. "
            "Do not just describe the code; provide the actual code block.\n"
            f"{context_prompt_part}" # Include file context if available
            f"Current Code in Editor:\n{code_content}\n\n"
            f"Instruction: {instruction}\n"
            "Response:\n"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw) # Process for charts
        return response_processed


class RAGFileBoxAgent: # Renamed from RAGDropboxAgent to RAGFileBoxAgent
    def __init__(self, chatbot: RAGChatBot): # Reference to chatbot to access doc_store
        self.chatbot = chatbot

    def run_instruction(self, instruction: str) -> str: # Instruction is directly from Dropbox tab NL box
        # Retrieve context from DocumentStore via chatbot
        relevant_chunks_with_names = self.chatbot.doc_store.retrieve_relevant_chunks(instruction, top_k=5) # Retrieve more context
        context_block = "\n\n".join([f"Relevant chunk from '{original_name}':\n{chunk}" for _, chunk, original_name in relevant_chunks_with_names])

        prompt = (
            "You are a highly capable and versatile AI assistant expert in analyzing and manipulating various file types within a File Box environment. " # Updated description
            "You have access to a wide range of tools and functionalities, including:\n"
            "   - Analyzing text, images, spreadsheets, and PDF documents.\n"
            "   - Summarizing information, extracting key insights, and answering questions based on file content.\n"
            "   - Generating code for data visualization and analysis using Python and matplotlib (indicate code blocks with triple backticks).\n"
            "   - Reasoning and problem-solving using the information available in the files.\n"
            "   - Accessing relevant context from emails and other files in the File Box to enrich your understanding and responses.\n" # Updated description
            "You should leverage these capabilities to understand user instructions related to files in File Box and provide comprehensive and helpful responses.\n" # Updated description
            "If the user asks to edit a text-based file, provide instructions on how to open it in the text editor and use natural language instructions within the editor for modification.\n"
            "Do not access external websites or local file system other than provided file context.\n\n"
            f"Context from File Box and Emails:\n{context_block}\n\n" # Updated description
            f"User Instruction: {instruction}\n"
            "Response:"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw)
        self.chatbot.retrieved_context_content = context_block # Update chatbot's context
        return response_processed


###############################################################################
# AI Code Editor (Jupyter-style - Enhanced UI)
###############################################################################
class AICodeEditor(ttk.Frame): # Inherits from Frame, not Toplevel
    def __init__(self, parent, master_app): # master_app reference
        super().__init__(parent)
        self.master_app = master_app # Store master app reference
        self.parent = parent # Store parent frame
        self.rag_code_agent = RAGCodeEditorAgent(self) # Agent for code editor
        self.code_namespace = {} # Namespace for code execution
        self.file_context_content = "" # To store combined content from relevant files

        # --- Main Paned Window for Vertical Split ---
        self.main_paned_window = ttk.Panedwindow(self, orient=tk.VERTICAL)
        self.main_paned_window.pack(fill="both", expand=True)

        # --- Top Paned Window for Code Input/Output (Horizontal Split) ---
        self.top_paned_window = ttk.Panedwindow(self.main_paned_window, orient=tk.HORIZONTAL)
        self.main_paned_window.add(self.top_paned_window, weight=1) # Weight for vertical resize

        # --- Code Input Frame ---
        self.input_frame = ttk.Frame(self.top_paned_window, relief=tk.SOLID, borderwidth=1) # Added border
        self.top_paned_window.add(self.input_frame, weight=2) # Weight for horizontal resize - wider input
        ttk.Label(self.input_frame, text="Code Input (Python):", font=("TkDefaultFont", 9, 'bold')).pack(anchor="w", padx=5, pady=(5,0)) # Bolder label
        self.code_input_area = ScrolledText(self.input_frame, wrap=tk.WORD, height=15, borderwidth=0) # Borderless inside frame
        self.code_input_area.pack(expand=True, fill='both', padx=5, pady=5)

        # --- Code Output Frame ---
        self.output_frame = ttk.Frame(self.top_paned_window, relief=tk.SOLID, borderwidth=1) # Added border
        self.top_paned_window.add(self.output_frame, weight=2) # Weight - equal output width
        ttk.Label(self.output_frame, text="Code Output/Visualization:", font=("TkDefaultFont", 9, 'bold')).pack(anchor="w", padx=5, pady=(5,0)) # Bolder label
        self.code_output_display = ScrolledText(self.output_frame, wrap=tk.WORD, height=10, state=tk.DISABLED, borderwidth=0) # Read-only for code output, borderless
        self.code_output_display.pack(expand=True, fill='both', padx=5, pady=5)

        # --- NL Instruction and Context Frame (BOTTOM - Vertical in Main Pane) ---
        self.bottom_frame = ttk.Frame(self.main_paned_window)
        self.main_paned_window.add(self.bottom_frame, weight=1) # Weight for vertical resize

        # ---- NL Instruction Input Frame ----
        nl_code_frame = ttk.LabelFrame(self.bottom_frame, text="Natural Language Instructions for Code Editor") # Frame in bottom frame
        nl_code_frame.pack(fill="both", expand=True, padx=10, pady=5)
        tk.Label(nl_code_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_code_instruction_entry = tk.Text(nl_code_frame, height=3, wrap=tk.WORD)
        self.nl_code_instruction_entry.pack(fill="x")
        self.nl_code_button = ttk.Button(nl_code_frame, text="Submit Code Instruction", command=self.process_code_nl_instruction)
        self.nl_code_button.pack(pady=5)

        # ---- Context Access Window (Below NL Instructions) ----
        self.context_code_frame = ttk.LabelFrame(self.bottom_frame, text="Context Accessed by Agent") # Frame for context
        self.context_code_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_code_display = ScrolledText(self.context_code_frame, height=8, wrap=tk.WORD, state=tk.DISABLED) # Even smaller context height
        self.context_code_display.pack(expand=True, fill='both')


        # --- Buttons in Input Frame (remains in input frame) ---
        button_input_frame = ttk.Frame(self.input_frame) # Buttons inside input frame
        button_input_frame.pack(fill="x", pady=5, padx=5) # Pad buttons from sides

        self.run_button = ttk.Button(button_input_frame, text="Run Code", command=self.run_code)
        self.run_button.pack(side="left", padx=5)
        self.clear_code_button = ttk.Button(button_input_frame, text="Clear Code", command=self.clear_code)
        self.clear_code_button.pack(side="left", padx=5)
        self.save_visual_button = ttk.Button(button_input_frame, text="Save Visualization", command=self.save_visualization, state=tk.DISABLED) # Initially disabled
        self.save_visual_button.pack(side="left", padx=5)
        self.load_file_context_button = ttk.Button(button_input_frame, text="Load File Context", command=self.load_file_context)
        self.load_file_context_button.pack(side="left", padx=5)


    def load_file_context(self): # ... (rest of AICodeEditor methods - no changes needed)
        # Get selected files from File Box and Gmail file lists in master_app
        dropbox_files = [self.master_app.filebox_file_listbox.get(idx) for idx in self.master_app.filebox_file_listbox.curselection()] # Updated to filebox_file_listbox
        gmail_messages = self.master_app.gmail_messages # Access stored gmail messages

        context_text = ""
        for filepath in dropbox_files:
            context_text += f"\n--- Content from File Box file: {os.path.basename(filepath)} ---\n" # Updated description
            context_text += extract_text_from_file(filepath)

        # Extract text from attachments of selected Gmail messages (you may need to refine this based on how you store/access attachments)
        if gmail_messages: # Assuming gmail_messages is a list of message dictionaries
            for msg in gmail_messages:
                if 'payload' in msg and 'parts' in msg['payload']:
                    for part in msg['payload']['parts']:
                        if 'filename' in part and part['filename']:
                            file_name = part['filename']
                            file_data_b64 = part['body']['data'] if 'body' in part and 'data' in part['body'] else None
                            if file_data_b64:
                                try:
                                    file_data = base64.urlsafe_b64decode(file_data_b64)
                                    temp_dir = tempfile.mkdtemp()
                                    file_path = os.path.join(temp_dir, file_name) # Temp file path
                                    with open(file_path, 'wb') as temp_file:
                                        temp_file.write(file_data)
                                    context_text += f"\n--- Content from Gmail attachment: {file_name} ---\n"
                                    context_text += extract_text_from_file(file_path) # Extract from temp file
                                except Exception as e:
                                    logging.error(f"Error processing Gmail attachment {file_name} for code editor context: {e}")

        self.file_context_content = context_text # Store combined context
        messagebox.showinfo("Context Loaded", "File context loaded for AI Code Editor. The AI agent can now access content from selected File Box and Gmail files when you give instructions.") # Updated message


    def process_code_nl_instruction(self): # ... (rest of AICodeEditor methods - no changes needed)
        instruction = self.nl_code_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return

        self.nl_code_instruction_entry.delete("1.0", tk.END)
        self.set_code_output("Processing code instruction...")
        self.set_context_code_output("") # Clear context

        threading.Thread(target=self._run_code_nl_instruction, args=(instruction,)).start()

    def _run_code_nl_instruction(self, instruction):
        code_content = self.code_input_area.get("1.0", tk.END)
        try:
            response = self.rag_code_agent.run_instruction(instruction, code_content, self.file_context_content) # Pass code and context
            self.set_code_output(response)
            if "[Visualization generated]" in response:
                self.enable_save_visual_button() # Enable save if visualization is generated
            else:
                self.disable_save_visual_button() # Otherwise, keep save disabled

            context_content = self.master_app.chatbot.get_last_retrieved_context() # Get context from chatbot
            self.set_context_code_output(context_content) # Display context in code editor


        except Exception as e:
            self.set_code_output(f"Error processing code instruction: {e}")
            self.set_context_code_output(f"Context retrieval error: {e}")
            self.disable_save_visual_button()


    def run_code(self):
        code_to_run = self.code_input_area.get("1.0", tk.END)
        self.set_code_output("Running code...")
        self.disable_save_visual_button() # Disable save button on each run

        threading.Thread(target=self._execute_code, args=(code_to_run,)).start()

    def _execute_code(self, code): # UPDATED - to display full output
        try:
            output = verify_and_run_code(code, self.code_namespace)
            self.set_code_output(output) # Output now includes stdout/stderr and visualization marker
            if "[Visualization Generated]" in output:
                self.enable_save_visual_button()
            else:
                self.disable_save_visual_button()

            context_content = self.master_app.chatbot.get_last_retrieved_context()
            self.set_context_code_output(context_content)

        except Exception as e:
            self.set_code_output(f"Error executing code: {e}")
            self.set_context_code_output(f"Context retrieval error: {e}")
            self.disable_save_visual_button()

    def set_code_output(self, text):
        self.code_output_display.config(state=tk.NORMAL) # Enable editing to insert
        self.code_output_display.delete("1.0", tk.END)
        self.code_output_display.insert("1.0", text)
        self.code_output_display.config(state=tk.DISABLED) # Disable editing

    def set_context_code_output(self, text): # New method for context display in code editor
        self.context_code_display.config(state=tk.NORMAL)
        self.context_code_display.delete("1.0", tk.END)
        self.context_code_display.insert("1.0", text)
        self.context_code_display.config(state=tk.DISABLED)

    def clear_code(self):
        self.code_input_area.delete("1.0", tk.END)
        self.code_output_display.config(state=tk.NORMAL)
        self.code_output_display.delete("1.0", tk.END)
        self.code_output_display.config(state=tk.DISABLED)
        self.set_context_code_output("") # Clear context as well
        self.disable_save_visual_button()

    def enable_save_visual_button(self):
        self.save_visual_button.config(state=tk.NORMAL)

    def disable_save_visual_button(self):
        self.save_visual_button.config(state=tk.DISABLED)

    def save_visualization(self): # ... (rest of AICodeEditor methods - no changes needed)
        fig = plt.gcf() # Get the current figure
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG files", "*.png"), ("All files", "*.*")])
        if file_path:
            try:
                fig.savefig(file_path)
                messagebox.showinfo("Visualization Saved", f"Visualization saved to {file_path}")
            except Exception as e:
                messagebox.showerror("Error Saving", f"Error saving visualization: {e}")


###############################################################################
# Main Application Class (MultiDocRAGApp) - Complete and Final
###############################################################################
class MultiDocRAGApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Multi-Document RAG Application")
        self.root.geometry("1400x900") # Adjusted for better layout
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        self.chatbot = RAGChatBot(DocumentStore(use_embeddings=True)) # Initialize RAG Chatbot and Document Store
        self.gmail_messages = [] # Store fetched gmail messages for context loading in code editor
        self.filebox_rag_agent = RAGFileBoxAgent(self.chatbot) # Renamed agent
        self.filebox_files_dir = "FileBox_files" # New directory for uploaded files
        os.makedirs(self.filebox_files_dir, exist_ok=True) # Ensure directory exists


        # --- Main Notebook ---
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        # --- RAG Chat Tab ---
        self.chat_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.chat_tab, text="RAG Chatbot")
        self.create_chat_tab_ui(self.chat_tab)

        # --- File Box Tab ---  // Renamed Tab
        self.filebox_tab = ttk.Frame(self.notebook) # Renamed tab Frame
        self.notebook.add(self.filebox_tab, text="File Box") # Renamed Tab Text
        self.create_filebox_tab_ui(self.filebox_tab) # Renamed UI function

        # --- Gmail Search Tab ---
        self.gmail_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.gmail_tab, text="Gmail Search")
        self.create_gmail_tab_ui(self.gmail_tab)

        # --- AI Code Editor Tab ---
        self.code_editor_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.code_editor_tab, text="AI Code Editor")
        self.create_code_editor_tab_ui(self.code_editor_tab)

        # --- Set Styles ---
        self.style = ttk.Style(root)
        self.style.theme_use('clam')  # 'clam', 'alt', 'default', 'classic', 'vista', 'xp', 'winnative'
        self.style.configure('TButton', padding=6, relief="flat", background="#f0f0f0")
        self.style.configure('TLabelFrame.Label', font=('TkDefaultFont', 10, 'bold'))

    ###########################################################################
    # --- RAG Chat Tab UI ---
    ###########################################################################
    def create_chat_tab_ui(self, parent_tab):
        parent_tab.grid_columnconfigure(0, weight=1)
        parent_tab.grid_rowconfigure(1, weight=1)

        # ---- File Loading Frame ----
        file_load_frame = ttk.LabelFrame(parent_tab, text="Document Loading")
        file_load_frame.grid(row=0, column=0, sticky='ew', padx=10, pady=5)

        self.file_listbox = tk.Listbox(file_load_frame, height=5, selectmode=tk.MULTIPLE)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        scrollbar_filelist = ttk.Scrollbar(file_load_frame, orient="vertical", command=self.file_listbox.yview)
        scrollbar_filelist.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_listbox.config(yscrollcommand=scrollbar_filelist.set)

        load_button = ttk.Button(file_load_frame, text="Load Local Files", command=self.load_local_files)
        load_button.pack(side=tk.BOTTOM, pady=5)

        # ---- Chat Frame ----
        chat_frame = ttk.LabelFrame(parent_tab, text="Chat Console")
        chat_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=5)
        chat_frame.grid_columnconfigure(0, weight=1)
        chat_frame.grid_rowconfigure(0, weight=1)

        self.chat_display_area = ScrolledText(chat_frame, wrap=tk.WORD, state=tk.DISABLED, height=15)
        self.chat_display_area.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        self.context_display_chat = ScrolledText(chat_frame, wrap=tk.WORD, state=tk.DISABLED, height=5) # Context display
        self.context_display_chat.grid(row=1, column=0, sticky='ew', padx=10, pady=5)
        tk.Label(chat_frame, text="Retrieved Context:").grid(row=1, column=0, sticky='w', padx=10) # Label for context

        input_frame = ttk.Frame(chat_frame)
        input_frame.grid(row=2, column=0, sticky='ew', padx=10, pady=10)
        self.chat_input_entry = ttk.Entry(input_frame)
        self.chat_input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.chat_input_entry.bind("<Return>", self.send_chat_message) # Bind Enter key
        send_button = ttk.Button(input_frame, text="Send", command=self.send_chat_message)
        send_button.pack(side=tk.RIGHT)

    def load_local_files(self): # ... (rest of MultiDocRAGApp methods - no changes needed)
        file_paths = filedialog.askopenfilenames(
            title="Choose files",
            filetypes=(("Text files", "*.txt"), ("PDF files", "*.pdf"), ("Excel files", "*.xlsx;*.xls"),
                       ("Word files", "*.docx"), ("CSV files", "*.csv"), ("Image files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"), ("All files", "*.*"))
        )
        if file_paths:
            for path in file_paths:
                file_name = os.path.basename(path)
                if file_name not in self.file_listbox.get(0, tk.END): # Prevent duplicates
                    self.file_listbox.insert(tk.END, file_name)
                    text_content = extract_text_from_file(path)
                    self.chatbot.add_document(text_content, file_name)
            messagebox.showinfo("Files Loaded", f"Loaded {len(file_paths)} files.")

    def send_chat_message(self, event=None): # ... (rest of MultiDocRAGApp methods - no changes needed)
        user_message = self.chat_input_entry.get()
        if user_message.strip():
            self.chat_input_entry.delete(0, tk.END)
            self.display_message("User:", user_message)
            self.display_message("Assistant:", "Thinking...")
            threading.Thread(target=self._get_chatbot_response, args=(user_message,)).start()

    def _get_chatbot_response(self, user_message): # ... (rest of MultiDocRAGApp methods - no changes needed)
        try:
            response = self.chatbot.chat(user_message)
            self.update_chat_display("Assistant:", response)
            context_content = self.chatbot.get_last_retrieved_context() # Get context
            self.update_context_display_chat(context_content) # Display context in chat tab
        except Exception as e:
            self.update_chat_display("Assistant:", f"Error: {e}")
            self.update_context_display_chat(f"Error retrieving context: {e}")

    def display_message(self, sender, message): # ... (rest of MultiDocRAGApp methods - no changes needed)
        self.chat_display_area.config(state=tk.NORMAL)
        self.chat_display_area.insert(tk.END, f"{sender} ", ('sender',))
        self.chat_display_area.insert(tk.END, f"{message}\n")
        self.chat_display_area.config(state=tk.DISABLED)
        self.chat_display_area.see(tk.END) # Autoscroll to end

    def update_chat_display(self, sender, message): # ... (rest of MultiDocRAGApp methods - no changes needed)
        self.chat_display_area.config(state=tk.NORMAL)
        #replace "Thinking..." message with actual response
        current_content = self.chat_display_area.get("1.0", tk.END)
        lines = current_content.splitlines()
        if lines and lines[-1] == "Assistant: Thinking...":
            self.chat_display_area.delete(f"{float(len(lines))} - 1 lines linestart", tk.END)

        self.chat_display_area.insert(tk.END, f"{sender} ", ('sender',))
        self.chat_display_area.insert(tk.END, f"{message}\n")
        self.chat_display_area.config(state=tk.DISABLED)
        self.chat_display_area.see(tk.END)

    def update_context_display_chat(self, context): # ... (rest of MultiDocRAGApp methods - no changes needed)
        self.context_display_chat.config(state=tk.NORMAL)
        self.context_display_chat.delete("1.0", tk.END)
        self.context_display_chat.insert("1.0", context)
        self.context_display_chat.config(state=tk.DISABLED)

    ###########################################################################
    # --- File Box Tab UI ---  // Renamed Tab and UI functions
    ###########################################################################
    def create_filebox_tab_ui(self, parent_tab): # Renamed UI function
        parent_tab.grid_columnconfigure(0, weight=1)
        parent_tab.grid_rowconfigure(1, weight=1)

        # ---- File Box File List Frame ---- // Renamed LabelFrame
        filebox_file_frame = ttk.LabelFrame(parent_tab, text="File Box Files") # Renamed LabelFrame
        filebox_file_frame.grid(row=0, column=0, sticky='ew', padx=10, pady=5)
        filebox_file_frame.grid_columnconfigure(0, weight=1)

        self.filebox_file_listbox = tk.Listbox(filebox_file_frame, height=5, selectmode=tk.MULTIPLE) # Renamed listbox
        self.filebox_file_listbox.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        scrollbar_filebox_list = ttk.Scrollbar(filebox_file_frame, orient="vertical", command=self.filebox_file_listbox.yview) # Renamed scrollbar
        scrollbar_filebox_list.grid(row=0, column=1, sticky='ns')
        self.filebox_file_listbox.config(yscrollcommand=scrollbar_filebox_list.set)

        # --- Buttons for File Box --- // Updated Buttons
        filebox_button_frame = ttk.Frame(filebox_file_frame) # Frame for buttons
        filebox_button_frame.grid(row=1, column=0, pady=5, columnspan=2) # Span columns

        upload_filebox_button = ttk.Button(filebox_button_frame, text="Upload Files to File Box", command=self.upload_files_to_filebox) # New Upload Button
        upload_filebox_button.pack(side=tk.LEFT, padx=5)
        load_initial_filebox_button = ttk.Button(filebox_button_frame, text="Load Default Files", command=self.load_initial_filebox_files) # Renamed Button and function
        load_initial_filebox_button.pack(side=tk.LEFT, padx=5)


        # ---- File Box RAG Interaction Frame ---- // Renamed LabelFrame
        filebox_rag_frame = ttk.LabelFrame(parent_tab, text="File Box File Interaction") # Renamed LabelFrame
        filebox_rag_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=5)
        filebox_rag_frame.grid_columnconfigure(0, weight=1)
        filebox_rag_frame.grid_rowconfigure(0, weight=1)

        self.filebox_instruction_area = ScrolledText(filebox_rag_frame, wrap=tk.WORD, height=10, state=tk.NORMAL) # Renamed instruction area
        self.filebox_instruction_area.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)

        self.context_display_filebox = ScrolledText(filebox_rag_frame, wrap=tk.WORD, state=tk.DISABLED, height=5) # Renamed context display
        self.context_display_filebox.grid(row=1, column=0, sticky='ew', padx=10, pady=5)
        tk.Label(filebox_rag_frame, text="Retrieved Context:").grid(row=1, column=0, sticky='w', padx=10) # Label

        filebox_nl_button = ttk.Button(filebox_rag_frame, text="Submit File Box Instruction", command=self.process_filebox_instruction) # Renamed button and function
        filebox_nl_button.grid(row=2, column=0, pady=10)

        open_filebox_button = ttk.Button(filebox_rag_frame, text="Open Selected File", command=self.open_selected_filebox_file) # Renamed button and function
        open_filebox_button.grid(row=3, column=0, pady=5)


    def load_initial_filebox_files(self): # Renamed function - loads default files
        initial_filebox_paths = [
            "Dropbox_files/sample_excel_file.xlsx",
            "Dropbox_files/sample_image.png",
            "Dropbox_files/sample_pdf.pdf",
            "Dropbox_files/sample_text_file.txt",
            "Dropbox_files/sample_word_file.docx",
            "Dropbox_files/state_of_union.txt",
        ] # Example File Box paths (using original Dropbox_files for simplicity)
        loaded_count = 0
        for path in initial_filebox_paths:
            if os.path.exists(path): # Check if file exists locally
                file_name = os.path.basename(path)
                if file_name not in self.filebox_file_listbox.get(0, tk.END): # Updated listbox name
                    self.filebox_file_listbox.insert(tk.END, file_name) # Updated listbox name
                    text_content = extract_text_from_file(path)
                    self.chatbot.add_document(text_content, file_name)
                    loaded_count += 1
        messagebox.showinfo("File Box Files Loaded", f"Loaded {loaded_count} default files into File Box.") # Updated message

    def upload_files_to_filebox(self): # New function - for uploading files
        file_paths = filedialog.askopenfilenames(
            title="Choose files to upload to File Box",
            filetypes=(("Text files", "*.txt"), ("PDF files", "*.pdf"), ("Excel files", "*.xlsx;*.xls"),
                       ("Word files", "*.docx"), ("CSV files", "*.csv"), ("Image files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"), ("All files", "*.*"))
        )
        if file_paths:
            uploaded_count = 0
            for file_path in file_paths:
                try:
                    file_name = os.path.basename(file_path)
                    dest_path = os.path.join(self.filebox_files_dir, file_name) # Destination in FileBox_files dir
                    shutil.copy(file_path, dest_path) # Copy file to FileBox_files dir
                    if file_name not in self.filebox_file_listbox.get(0, tk.END): # Check for duplicates in listbox
                        self.filebox_file_listbox.insert(tk.END, file_name) # Add to listbox
                        text_content = extract_text_from_file(dest_path) # Extract from copied file
                        self.chatbot.add_document(text_content, file_name) # Add to document store
                        uploaded_count += 1
                except Exception as e:
                    logging.error(f"Error uploading file {file_name} to File Box: {e}")
                    messagebox.showerror("Upload Error", f"Error uploading {file_name}: {e}")
            if uploaded_count > 0:
                messagebox.showinfo("Upload Successful", f"Successfully uploaded {uploaded_count} files to File Box.") # Updated message


    def process_filebox_instruction(self): # Renamed function
        instruction = self.filebox_instruction_area.get("1.0", tk.END).strip() # Updated area name
        if instruction:
            self.display_filebox_response("Assistant:", "Thinking...") # Updated function name
            threading.Thread(target=self._get_filebox_response, args=(instruction,)).start() # Updated function name

    def _get_filebox_response(self, instruction): # Renamed function
        try:
            response = self.filebox_rag_agent.run_instruction(instruction) # Use FileBox agent
            self.update_filebox_display("Assistant:", response) # Updated display function name
            context_content = self.chatbot.get_last_retrieved_context() # Get context
            self.update_context_display_filebox(context_content) # Updated context display function name
        except Exception as e:
            self.update_filebox_display("Assistant:", f"Error: {e}") # Updated display function name
            self.update_context_display_filebox(f"Error retrieving context: {e}") # Updated context display function name


    def display_filebox_response(self, sender, message): # Renamed function
        self.filebox_instruction_area.config(state=tk.NORMAL) # Updated area name
        self.filebox_instruction_area.insert(tk.END, f"\n{sender} {message}\n", ('sender',)) # Updated area name
        self.filebox_instruction_area.config(state=tk.DISABLED) # Re-disable
        self.filebox_instruction_area.see(tk.END)

    def update_filebox_display(self, sender, message): # Renamed function
        self.filebox_instruction_area.config(state=tk.NORMAL) # Updated area name
        current_content = self.filebox_instruction_area.get("1.0", tk.END) # Updated area name
        lines = current_content.splitlines()
        if lines and lines[-1] == "Assistant: Thinking...":
            self.filebox_instruction_area.delete(f"{float(len(lines))} - 1 lines linestart", tk.END) # Updated area name

        self.filebox_instruction_area.insert(tk.END, f"\n{sender} {message}\n", ('sender',)) # Updated area name
        self.filebox_instruction_area.config(state=tk.DISABLED)
        self.filebox_instruction_area.see(tk.END)

    def update_context_display_filebox(self, context): # Renamed function
        self.context_display_filebox.config(state=tk.NORMAL) # Updated area name
        self.context_display_filebox.delete("1.0", tk.END) # Updated area name
        self.context_display_filebox.insert("1.0", context) # Updated area name
        self.context_display_filebox.config(state=tk.DISABLED)

    def open_selected_filebox_file(self): # Renamed function
        selected_files = self.filebox_file_listbox.curselection() # Updated listbox name
        if not selected_files:
            messagebox.showinfo("Info", "Please select a file to open.")
            return
        file_name = self.filebox_file_listbox.get(selected_files[0]) # Updated listbox name
        file_path = os.path.join(self.filebox_files_dir, file_name) # Updated dir name
        if os.path.exists(file_path):
            open_custom_editor(self.root, file_path, self) # Pass master_app


    ###########################################################################
    # --- Gmail Tab UI ---
    ###########################################################################
    def create_gmail_tab_ui(self, parent_tab): # ... (rest of MultiDocRAGApp methods - no changes needed)
        parent_tab.grid_columnconfigure(0, weight=1)
        parent_tab.grid_rowconfigure(2, weight=1) # Results area row to expand

        # ---- Gmail Query Frame ----
        gmail_query_frame = ttk.Frame(parent_tab)
        gmail_query_frame.grid(row=0, column=0, sticky='ew', padx=10, pady=10)
        tk.Label(gmail_query_frame, text="Enter Gmail Search Query:").pack(side=tk.LEFT, padx=5)
        self.gmail_query_entry = ttk.Entry(gmail_query_frame)
        self.gmail_query_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.gmail_search_button = ttk.Button(gmail_query_frame, text="Search Gmail", command=self.search_gmail)
        self.gmail_search_button.pack(side=tk.LEFT, padx=5)

        # ---- Gmail Results Display Frame ----
        self.gmail_results_frame = ttk.LabelFrame(parent_tab, text="Gmail Search Results")
        self.gmail_results_frame.grid(row=2, column=0, sticky='nsew', padx=10, pady=10)
        self.gmail_results_area = ScrollableFrame(self.gmail_results_frame) # Use ScrollableFrame
        self.gmail_results_area.pack(fill="both", expand=True, padx=10, pady=10)

        # ---- Fetch Mode Frame ----
        gmail_fetch_mode_frame = ttk.Frame(parent_tab)
        gmail_fetch_mode_frame.grid(row=1, column=0, sticky='ew', padx=10, pady=5)
        self.fetch_mode_var = tk.StringVar(value='summary') # Default fetch mode
        tk.Radiobutton(gmail_fetch_mode_frame, text="Summary", variable=self.fetch_mode_var, value='summary').pack(side=tk.LEFT)
        tk.Radiobutton(gmail_fetch_mode_frame, text="Full", variable=self.fetch_mode_var, value='full').pack(side=tk.LEFT)
        tk.Radiobutton(gmail_fetch_mode_frame, text="Analyze", variable=self.fetch_mode_var, value='analyze').pack(side=tk.LEFT)
        self.gmail_status_label = tk.Label(gmail_fetch_mode_frame, text="") # Status label
        self.gmail_status_label.pack(side=tk.LEFT, padx=10)


    def search_gmail(self): # ... (rest of MultiDocRAGApp methods - no changes needed)
        query = self.gmail_query_entry.get()
        if not query.strip():
            query = "in:inbox" # Default to inbox if query is empty
        fetch_mode = self.fetch_mode_var.get()
        self.gmail_status_label.config(text="Searching Gmail...")
        self.gmail_messages = [] # Clear previous messages before new search
        threading.Thread(target=self._perform_gmail_search, args=(query, fetch_mode)).start()

    def _perform_gmail_search(self, query, fetch_mode): # ... (rest of MultiDocRAGApp methods - no changes needed)
        try:
            interpreted_query_data = interpret_user_query(query)
            gmail_query = interpreted_query_data.get("gmail_query", query) # Use interpreted query or original if interpretation fails
            has_attachments = interpreted_query_data.get("has_attachments", False) # Get has_attachments flag
            messages = fetch_gmail_data(gmail_query, batch_size=10) # Reduced batch size for UI demo

            if not messages:
                self.gmail_status_label.config(text="No emails found.")
                self._display_email_summaries([], fetch_mode, has_attachments) # Display empty results
                return

            self.gmail_messages = messages # Store fetched messages
            self.gmail_status_label.config(text=f"Found {len(messages)} emails. Displaying summaries...")
            self._display_email_summaries(messages, fetch_mode, has_attachments) # Pass has_attachments flag

        except Exception as e:
            logging.error(f"Gmail search error: {e}")
            self.gmail_status_label.config(text=f"Gmail Search Error: {e}")
            self._display_email_summaries([], fetch_mode, False) # Display empty results on error


    def _display_email_summaries(self, messages, fetch_mode, has_attachments): # UPDATED - with functools.partial and logging
        # Clear previous results
        for widget in self.gmail_results_area.scrollable_frame.winfo_children(): # Access scrollable_frame
            widget.destroy()

        if not messages:
            tk.Label(self.gmail_results_area.scrollable_frame, text="No emails found matching your query.").pack(padx=10, pady=10) # Use scrollable_frame
            return

        for msg in messages:
            msg_frame = ttk.Frame(self.gmail_results_area.scrollable_frame, padding=10, borderwidth=2, relief=tk.GROOVE) # Use scrollable_frame
            msg_frame.pack(pady=5, fill="x")

            headers = dict([(h['name'], h['value']) for h in msg['payload']['headers']])
            subject = headers.get('Subject', 'No Subject')
            sender = headers.get('From', 'Unknown Sender')
            date = headers.get('Date', 'Unknown Date')

            tk.Label(msg_frame, text=f"Subject: {subject}", anchor="w", font=("TkDefaultFont", 10, 'bold')).pack(fill="x")
            tk.Label(msg_frame, text=f"From: {sender}", anchor="w").pack(fill="x")
            tk.Label(msg_frame, text=f"Date: {date}", anchor="w").pack(fill="x")

            summary_text = self._summarize_email_content(msg, fetch_mode)
            tk.Label(msg_frame, text="Summary:").pack(anchor="w")
            tk.Label(msg_frame, text=summary_text, anchor="w", wraplength=500).pack(fill="x")

            # --- Buttons ---
            button_frame = ttk.Frame(msg_frame)
            button_frame.pack(fill="x", pady=5)

            open_email_btn = ttk.Button(button_frame, text="Open Email", command=lambda m=msg: self._open_email_content(m)) # Pass msg
            open_email_btn.pack(side="left", padx=5)

            # --- Attachment Handling ---
            if 'parts' in msg['payload']:
                attachments_frame = ttk.Frame(button_frame)
                attachments_frame.pack(side="left", padx=5)
                has_any_attachment = False # Flag to check if any attachment was found

                for part in msg['payload']['parts']:
                    if 'filename' in part and part['filename']:
                        file_name = part['filename']
                        has_any_attachment = True # Set flag if attachment is found
                        # --- FIX: Use functools.partial to fix lambda closure ---
                        cmd = functools.partial(self._open_attachment, part=part, msg=msg) # Create partial function
                        open_file_btn = ttk.Button(attachments_frame, text=file_name, command=cmd) # Use partial function
                        # --- Debug Print ---
                        print(f"Button command created for attachment: {file_name}, part ID: {part.get('partId') if isinstance(part, dict) else 'N/A'}, msg ID: {msg.get('id')}")
                        open_file_btn.pack(side="left", padx=2)

                if has_any_attachment: # Only show "Attachments:" label if there are attachments
                    tk.Label(attachments_frame, text="Attachments:").pack(side="left")
        self.gmail_status_label.config(text=f"Displayed {len(messages)} emails.") # Update status at the end


    def _summarize_email_content(self, msg, fetch_mode): # ... (rest of MultiDocRAGApp methods - no changes needed)
        if fetch_mode == 'summary':
            plain_text_body = ""
            html_body = ""

            if msg['payload'].get('parts'): # Check for parts for multipart messages
                for part in msg['payload']['parts']:
                    if part['mimeType'] == 'text/plain':
                        plain_text_body = base64.urlsafe_b64decode(part['body']['data']).decode() if 'data' in part['body'] else ""
                        break # Prioritize plain text
                    elif part['mimeType'] == 'text/html' and not plain_text_body: # Fallback to HTML if no plain text
                        html_body = base64.urlsafe_b64decode(part['body']['data']).decode() if 'data' in part['body'] else ""

            elif msg['payload']['mimeType'] == 'text/plain': # For non-multipart, plain text messages
                plain_text_body = base64.urlsafe_b64decode(msg['payload']['body']['data']).decode() if 'data' in msg['payload']['body'] else ""
            elif msg['payload']['mimeType'] == 'text/html': # Fallback to HTML for non-multipart
                html_body = base64.urlsafe_b64decode(msg['payload']['body']['data']).decode() if 'data' in msg['payload']['body'] else ""

            email_content = plain_text_body or html_body # Use plain text if available, else HTML

            if not email_content: # If still no content, indicate no body found
                return "[No email body content found for summary]"


            prompt = (
                "Summarize the following email content:\n"
                f"{email_content}\n"
                "Summary:"
            )
            return llm_predict_with_backoff(prompt)
        elif fetch_mode in ['full', 'analyze']: # For 'full' and 'analyze' modes, return full content
            plain_text_body = ""
            html_body = ""

            if msg['payload'].get('parts'): # Check for parts for multipart messages
                for part in msg['payload']['parts']:
                    if part['mimeType'] == 'text/plain':
                        plain_text_body = base64.urlsafe_b64decode(part['body']['data']).decode() if 'data' in part['body'] else ""
                        break # Prioritize plain text
                    elif part['mimeType'] == 'text/html' and not plain_text_body: # Fallback to HTML if no plain text
                        html_body = base64.urlsafe_b64decode(part['body']['data']).decode() if 'data' in msg['payload']['body'] else ""
            elif msg['payload']['mimeType'] == 'text/plain': # For non-multipart, plain text messages
                plain_text_body = base64.urlsafe_b64decode(msg['payload']['body']['data']).decode() if 'data' in msg['payload']['body'] else ""
            elif msg['payload']['mimeType'] == 'text/html': # Fallback to HTML for non-multipart
                html_body = base64.urlsafe_b64decode(msg['payload']['body']['data']).decode() if 'data' in msg['payload']['body'] else ""

            email_content = plain_text_body or html_body

            if not email_content:
                return "[No email body content found for full/analyze mode]"
            return email_content
        return "[Unknown fetch mode]"

    def _open_email_content(self, msg): # ... (rest of MultiDocRAGApp methods - no changes needed)
        fetch_mode = self.fetch_mode_var.get()
        email_content = self._summarize_email_content(msg, fetch_mode) # Fetch content again, respect fetch mode

        email_win = tk.Toplevel(self.root)
        email_win.title(f"Email Content - Subject: {dict([(h['name'], h['value']) for h in msg['payload']['headers']]).get('Subject', 'No Subject')}")
        email_win.geometry("900x700")
        text_area = ScrolledText(email_win, wrap=tk.WORD)
        text_area.pack(expand=True, fill='both', padx=10, pady=10)
        text_area.insert('1.0', email_content)
        text_area.config(state=tk.DISABLED)


    def _open_attachment(self, part, msg): # UPDATED - with enhanced logging and error handling
        file_name = part['filename']
        logging.info(f"_open_attachment called for: {file_name}, part ID: {part.get('partId') if isinstance(part, dict) else 'N/A'}, msg ID: {msg.get('id')}") # Log function call
        try:
            file_data_b64 = part['body']['data']
            if not file_data_b64:
                logging.warning(f"No data found in part body for {file_name}, part ID: {part.get('partId') if isinstance(part, dict) else 'N/A'}, msg ID: {msg.get('id')}")
                messagebox.showerror("Attachment Error", f"No data found for attachment: {file_name}")
                return # Exit if no data

            file_data = base64.urlsafe_b64decode(file_data_b64)
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, file_name)

            logging.info(f"Saving attachment {file_name} to temp file: {file_path}") # Log temp file path
            with open(file_path, 'wb') as temp_file:
                temp_file.write(file_data)

            logging.info(f"Opening custom editor for: {file_path}") # Log before opening editor
            open_custom_editor(self.root, file_path, self) # Pass master_app
            logging.info(f"Custom editor opened for: {file_path} successfully.") # Log after editor opens

        except Exception as e:
            messagebox.showerror("Attachment Error", f"Error opening attachment {file_name}: {e}")
            logging.error(f"Error opening attachment {file_name}: {e}", exc_info=True) # Log full exception info


    ###########################################################################
    # --- AI Code Editor Tab UI ---
    ###########################################################################
    def create_code_editor_tab_ui(self, parent_tab):
        code_editor = AICodeEditor(parent_tab, master_app=self) # Pass master_app instance
        code_editor.pack(fill="both", expand=True, padx=10, pady=10)


###############################################################################
# Main execution
###############################################################################
def main():
    root = tk.Tk()
    app = MultiDocRAGApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()