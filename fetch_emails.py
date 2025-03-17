from __future__ import print_function
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pickle
import os.path
import base64

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

def fetch_gmail_data(query=None, download_attachments=True):
    """
    Fetches all emails and attachments from Gmail based on a query.

    Args:
        query: Gmail search query (e.g., "subject:meeting notes"). If None, fetches all emails.
        download_attachments: Whether to download attachments.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('gmail', 'v1', credentials=creds)

    if query:
        # Call the Gmail API to search for emails based on a query
        results = service.users().messages().list(userId='me', q=query).execute()
    else:
        # Call the Gmail API to fetch inbox
        results = service.users().messages().list(userId='me', labelIds=['INBOX']).execute()

    messages = results.get('messages', [])

    for message in messages:
        msg = service.users().messages().get(userId='me', id=message['id'], format='full').execute()

        # Process message details here
        print(f"Message ID: {message['id']}")
        print(f"Snippet: {msg['snippet']}")

        # Download attachments if requested
        if download_attachments:
            payload = msg['payload']
            if 'parts' in payload:
                for part in payload['parts']:
                    if 'filename' in part and part['filename']:
                        file_name = part['filename']
                        mime_type = part['mimeType']
                        print(f"Attachment Filename: {file_name}")

                        if 'data' in part['body']:
                            data = part['body']['data']
                        else:
                            attachment_id = part['body']['attachmentId']
                            attachment = service.users().messages().attachments().get(
                                userId='me', messageId=message['id'], id=attachment_id
                            ).execute()
                            data = attachment['data']

                        file_data = base64.urlsafe_b64decode(data.encode('UTF-8'))
                        file_path = os.path.join('.', file_name)  # Save in current directory
                        with open(file_path, 'wb') as f:
                            f.write(file_data)
                        print(f"Attachment '{file_name}' downloaded to '{file_path}'")

# Example usage with a query and attachment downloading
fetch_gmail_data(download_attachments=True)













#!/usr/bin/env python3
from __future__ import print_function
import os
import re
import json
import pickle
import base64
import io
import statistics
import time
import logging
import csv
import threading
import tempfile
import queue

from dotenv import load_dotenv
from typing import Optional, Any, List, Dict, ClassVar

from pydantic import Field, BaseModel

from crewai import Agent, Task, Crew, LLM, Process
from crewai.tools.base_tool import BaseTool

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

from PyPDF2 import PdfReader
import openpyxl
import docx

from litellm.exceptions import RateLimitError

try:
    from PIL import Image
    import pytesseract
except ImportError:
    pytesseract = None

import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from tkinter import filedialog
from tkinter import messagebox
try:
    from tkcalendar import DateEntry
except ImportError:
    print("tkcalendar not installed. Date selection will be unavailable.")
    DateEntry = None

from concurrent.futures import ThreadPoolExecutor
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# Load environment variables and configure API keys/scopes
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

# Initialize the LLM (Gemini) using Crew AI's LLM class.
gemini_llm = LLM(
    model="gemini/gemini-2.0-flash",
    api_key=GEMINI_API_KEY,
    temperature=0.1
)

# Configure Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Helper: LLM call with exponential backoff + throttle
def llm_predict_with_backoff(prompt: str, max_retries: int = 5, initial_delay: float = 1.0) -> str:
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            time.sleep(2.0)
            response = gemini_llm.call(prompt)
            return response
        except RateLimitError as e:
            logging.warning("Rate limit error on attempt %d/%d: %s", attempt + 1, max_retries, e)
            time.sleep(delay)
            delay *= 2
        except Exception as e:
            logging.error("Error during LLM call: %s", e)
            return ""
    raise RateLimitError("Exceeded maximum retry attempts for LLM call")

# Standalone Gmail Fetching Function with Batching
def fetch_gmail_data(query: Optional[str] = None, batch_size: int = 50) -> List[Dict[str, Any]]:
    creds = None
    token_path = 'token.pickle'
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'wb') as token:
            pickle.dump(creds, token)

    service = build('gmail', 'v1', credentials=creds)
    fetched_messages = []
    page_token = None

    while len(fetched_messages) < batch_size:
        if query:
            results = service.users().messages().list(
                userId='me',
                q=query,
                maxResults=batch_size,
                pageToken=page_token
            ).execute()
        else:
            results = service.users().messages().list(
                userId='me',
                labelIds=['INBOX'],
                maxResults=batch_size,
                pageToken=page_token
            ).execute()
        messages = results.get('messages', [])
        page_token = results.get('nextPageToken')
        for message in messages:
            msg = service.users().messages().get(
                userId='me',
                id=message['id'],
                format='full'
            ).execute()
            snippet = msg.get('snippet', 'No snippet available')
            short_snippet = (snippet[:200] + '...') if len(snippet) > 200 else snippet
            logging.info("Fetched Message ID: %s - Snippet: %s", message['id'], short_snippet)
            fetched_messages.append(msg)
            if len(fetched_messages) >= batch_size:
                break
        if not page_token:
            break
        time.sleep(1)
    return fetched_messages

# Crew AI Tool: GmailDataFetcherTool
class GmailDataFetcherTool(BaseTool):
    name: str = "gmail_data_fetcher"
    description: str = (
        "Fetches Gmail data and attachments, processes emails using an LLM, supports follow-up queries, and "
        "can store extracted data upon request. Also supports analysis of Excel, CSV, PDF, image, and PowerBI files "
        "with custom chunking and downloading attachments."
    )
    config: Dict[str, Any] = Field(default_factory=dict)
    conversation_history: List[str] = Field(default_factory=list)
    cache: ClassVar[Dict[str, Any]] = Field(default_factory=dict)

    def __init__(self, **data: Any):
        super().__init__(**data)
        object.__setattr__(self, "config", self.load_external_config("search_config.json"))

    def load_external_config(self, filename: str) -> Dict[str, Any]:
        if os.path.exists(filename):
            with open(filename, "r") as file:
                return json.load(file)
        return {}

    def advanced_query_parsing(self, query: str) -> str:
        tokens = re.split(r'\W+', query)
        stopwords = self.config.get("stopwords", [])
        filtered_tokens = [token for token in tokens if token and token.lower() not in stopwords]
        synonyms = self.config.get("synonyms", {})
        normalized_tokens = [synonyms.get(token.lower(), token) for token in filtered_tokens]
        parsed = " ".join(normalized_tokens)
        logging.info("Advanced parsed query: %s", parsed)
        return parsed

    def perform_calculation(self, numbers: List[float], operation: str) -> Optional[float]:
        if not numbers:
            return None
        if operation == "addition":
            return sum(numbers)
        elif operation == "subtraction":
            result = numbers[0]
            for num in numbers[1:]:
                result -= num
            return result
        elif operation == "average":
            return sum(numbers) / len(numbers)
        elif operation == "median":
            return statistics.median(numbers)
        elif operation == "stdev":
            return statistics.stdev(numbers) if len(numbers) > 1 else 0.0
        else:
            return None

    def group_data(self, data: List[Dict[str, Any]], group_by: str) -> Dict[Any, List[Dict[str, Any]]]:
        grouped = {}
        for item in data:
            key = item.get(group_by, "Undefined")
            grouped.setdefault(key, []).append(item)
        return grouped

    def get_auto_complete_suggestions(self, partial_query: str) -> List[str]:
        suggestions = [
            partial_query + " report",
            partial_query + " summary",
            partial_query + " analysis"
        ]
        logging.info("Auto-complete suggestions: %s", suggestions)
        return suggestions

    def chunk_text(self, text: str, method: str) -> List[str]:
        if method == "word":
            words = text.split()
            return [" ".join(words[i:i + 50]) for i in range(0, len(words), 50)]
        elif method == "sentence":
            sentences = re.split(r'(?<=[.!?])\s+', text)
            return sentences
        elif method == "paragraph":
            paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
            return paragraphs
        elif method == "page":
            pages = text.split("\f")
            return pages
        elif method == "topic":
            topics = [topic.strip() for topic in text.split('\n\n') if topic.strip()]
            return topics if topics else [text]
        else:
            return [text]

    def parse_chunking_method(self, query: str) -> Optional[str]:
        q = query.lower()
        if "chunk by word" in q:
            return "word"
        elif "chunk by sentence" in q:
            return "sentence"
        elif "chunk by paragraph" in q:
            return "paragraph"
        elif "chunk by page" in q:
            return "page"
        elif "chunk by topic" in q:
            return "topic"
        else:
            return None

    def extract_email_text(self, msg: Dict[str, Any], chunk_method: Optional[str] = None,
                             download_attachments: bool = False, download_path: Optional[str] = None, analyze_files: bool = False) -> str:
        try:
            email_text = ""
            payload = msg.get('payload', {})
            if 'parts' in payload:
                for part in payload['parts']:
                    mime_type = part.get('mimeType', '')
                    if mime_type == 'text/plain':
                        data = part['body'].get('data')
                        if data:
                            email_text += base64.urlsafe_b64decode(data.encode('UTF-8')).decode('UTF-8')
                    elif mime_type == 'text/html':
                        pass
                    if part.get('filename'):
                        file_name = part.get('filename')
                        email_text += "\n\nAttachment: " + file_name + "\n"
                        if analyze_files:
                            analysis_result = self.analyze_attachment(msg['id'], part)
                            email_text += analysis_result
                        else:
                            attachment_text = self.get_attachment_text(msg['id'], part, chunk_method)
                            email_text += attachment_text

                        if download_attachments:
                            self.download_attachment(msg['id'], part, download_path)
            else:
                if payload.get('mimeType') == 'text/plain':
                    data = payload['body'].get('data')
                    if data:
                        email_text = base64.urlsafe_b64decode(data.encode('UTF-8')).decode('UTF-8')
            return email_text
        except Exception as e:
            logging.error("Error extracting email text: %s", e)
            return f"Error extracting email text: {e}"

    def get_attachment_text(self, message_id: str, part: Dict[str, Any], chunk_method: Optional[str] = None) -> str:
        try:
            file_name = part.get('filename', 'unknown')
            temp_dir = tempfile.mkdtemp()
            file_path = os.path.join(temp_dir, file_name)
            self.download_attachment(message_id, part, temp_dir)

            text = self.extract_text_from_file(file_path, file_name)

            if text and chunk_method:
                chunks = self.chunk_text(text, chunk_method)
                text = "\n--- Chunk ---\n".join(chunks)

            self.cleanup_temp_dir(temp_dir)
            return text
        except Exception as e:
            logging.error("Error extracting text from %s: %s", file_name, e)
            return f"Error extracting text from {file_name}: {e}"

    def extract_text_from_file(self, file_path: str, file_name: str) -> str:
        text = ""
        fname = file_name.lower()
        try:
            if fname.endswith('.pdf'):
                with open(file_path, 'rb') as file_content:
                    reader = PdfReader(file_content)
                    text = "".join(p.extract_text() for p in reader.pages if p.extract_text())
            elif fname.endswith(('.xlsx', '.xls')):
                with open(file_path, 'rb') as file_content:
                    workbook = openpyxl.load_workbook(file_content)
                    for sheet in workbook.sheetnames:
                        worksheet = workbook[sheet]
                        for row in worksheet.iter_rows():
                            text += " ".join(str(cell.value) for cell in row if cell.value is not None) + "\n"
            elif fname.endswith('.docx'):
                with open(file_path, 'rb') as file_content:
                    document = docx.Document(file_content)
                    text = "\n".join(p.text for p in document.paragraphs)
            elif fname.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as file_content:
                    reader = csv.reader(file_content)
                    text = "\n".join([",".join(row) for row in reader])
            elif fname.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                if pytesseract:
                    try:
                        image = Image.open(file_path)
                        text = pytesseract.image_to_string(image)
                    except Exception as e:
                        text = f"Error during OCR: {e}"
                else:
                    text = "OCR support not installed for image extraction."
            elif fname.endswith('.pbix'):
                text = "PowerBI (.pbix) file extraction is not directly supported."
            return text
        except Exception as e:
            logging.error("Error extracting text from %s: %s", file_name, e)
            return f"Error extracting text from {file_name}: {e}"

    def download_attachment(self, message_id: str, part: Dict[str, Any], download_path: str):
        try:
            service = build('gmail', 'v1', credentials=self.get_gmail_credentials())
            attachment_id = part['body']['attachmentId']
            response = service.users().messages().attachments().get(
                userId='me',
                messageId=message_id,
                id=attachment_id
            ).execute()
            file_data = base64.urlsafe_b64decode(response['data'].encode('UTF-8'))
            file_path = os.path.join(download_path, part['filename'])
            with open(file_path, 'wb') as file:
                file.write(file_data)
            logging.info("Attachment downloaded to %s", file_path)
        except Exception as e:
            logging.error("Error downloading attachment: %s", e)
            return f"Error downloading attachment: {e}"

    def get_gmail_credentials(self):
        creds = None
        token_path = 'token.pickle'
        if os.path.exists(token_path):
            with open(token_path, 'rb') as token:
                creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_path, 'wb') as token:
                pickle.dump(creds, token)
        return creds

    def cleanup_temp_dir(self, temp_dir: str):
        import shutil
        shutil.rmtree(temp_dir)

    def analyze_attachment(self, message_id: str, part: Dict[str, Any]) -> str:
        # Placeholder for custom analysis logic
        return "Attachment analysis placeholder"

    def _run(self, query: str) -> str:
        """The entrypoint for the tool."""
        try:
            logging.info(f"Running GmailDataFetcherTool with query: {query}")
            # 1. Parse the query
            parsed_query = self.advanced_query_parsing(query)
            logging.info(f"Parsed query: {parsed_query}")

            # 2. Determine if chunking is needed
            chunk_method = self.parse_chunking_method(query)
            logging.info(f"Chunk method: {chunk_method}")

            # 3. Determine download needs
            download_attachments = "download attachments" in query.lower()
            download_path = None
            if download_attachments:
                download_path = filedialog.askdirectory(title="Select Directory to Save Attachments")
                if not download_path:
                    return "Attachment download cancelled."
                logging.info(f"Attachments will be downloaded to: {download_path}")

            analyze_files = "analyze files" in query.lower()

            # 4. Fetch Gmail data
            messages = fetch_gmail_data(query=parsed_query, batch_size=10)  # Reduced batch size for testing
            logging.info(f"Fetched {len(messages)} messages.")

            # 5. Extract email text
            all_email_text = ""
            for msg in messages:
                email_text = self.extract_email_text(msg, chunk_method, download_attachments, download_path, analyze_files)
                all_email_text += email_text + "\n---\n"

            # 6. Handle post-processing operations (calculations, grouping, etc.)

            # 7. Respond with LLM or return directly
            # This is a placeholder: Adjust logic based on tool's configuration.
            if "summary" in query.lower() or "report" in query.lower():
                prompt = f"Summarize the following email data:\n{all_email_text}"
                llm_response = llm_predict_with_backoff(prompt)
                return llm_response
            else:
                return all_email_text

        except Exception as e:
            logging.error(f"Tool execution failed: {e}", exc_info=True)
            return f"Error processing Gmail data: {e}"

# Simple UI for Gmail Data Fetcher
class GmailDataFetcherUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Gmail Data Fetcher")
        self.fetcher_tool = GmailDataFetcherTool()

        # Input frame
        self.input_frame = ttk.Frame(self.root)
        self.input_frame.pack(fill="x")

        self.query_label = ttk.Label(self.input_frame, text="Enter Gmail Query:")
        self.query_label.pack(side="left")

        self.query_entry = ttk.Entry(self.input_frame, width=50)
        self.query_entry.pack(side="left", padx=5)

        self.fetch_button = ttk.Button(self.input_frame, text="Fetch and Process", command=self.fetch_and_process)
        self.fetch_button.pack(side="left")

        # Output area
        self.output_text = ScrolledText(self.root, wrap=tk.WORD, height=20)
        self.output_text.pack(fill="both", expand=True, padx=10, pady=10)

        # Status bar
        self.status_bar = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill="x")

    def fetch_and_process(self):
        query = self.query_entry.get()
        if not query:
            messagebox.showerror("Error", "Query cannot be empty.")
            return

        self.status_bar.config(text="Fetching and processing...")
        self.root.update_idletasks()

        try:
            # Execute the tool in a separate thread to prevent UI blocking
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(self.fetcher_tool._run, query)
                result = future.result()

            self.output_text.delete("1.0", tk.END)
            self.output_text.insert("1.0", result)
            self.status_bar.config(text="Done.")

        except Exception as e:
            self.output_text.delete("1.0", tk.END)
            self.output_text.insert("1.0", f"Error: {e}")
            self.status_bar.config(text=f"Error: {e}")
        finally:
            self.root.update_idletasks()

if __name__ == "__main__":
    root = tk.Tk()
    ui = GmailDataFetcherUI(root)
    root.mainloop()
