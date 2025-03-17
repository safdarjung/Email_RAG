import os
import re
import json
import pickle
import base64
import csv
import time
import logging
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinter.scrolledtext import ScrolledText
from typing import Any, Dict, List, Optional, Tuple, ClassVar
import contextlib
import io
import functools
import shutil
import pandas as pd

from dotenv import load_dotenv
from pydantic import Field
from crewai import LLM, Agent
from crewai.tools.base_tool import BaseTool

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials

from PyPDF2 import PdfReader
import openpyxl
import docx

try:
    from PIL import Image, ImageTk
    import pytesseract
except ImportError:
    pytesseract = None

try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib

matplotlib.use('TkAgg')

try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None

try:
    from sentence_transformers import SentenceTransformer, util
    EMBEDDING_MODEL = SentenceTransformer('all-MiniLM-L6-v2')
except ImportError:
    EMBEDDING_MODEL = None

import torch

# Load environment variables and set scopes
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "default_value")
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', filename='app.log', filemode='w')

###############################################################################
# Global LLM instance
###############################################################################
unified_llm = LLM(
    model="gemini/gemini-2.0-flash",
    api_key=GEMINI_API_KEY,
    temperature=0.1
)

###############################################################################
# Utility Functions
###############################################################################
def llm_predict_with_backoff(prompt: str, max_retries: int = 5, initial_delay: float = 1.0) -> str:
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            time.sleep(2.0)
            response = unified_llm.call(prompt)
            return response
        except Exception as e:
            logging.warning("Rate limit error on attempt %d/%d: %s", attempt+1, max_retries, e)
            time.sleep(delay)
            delay *= 2
        except Exception as e:
            logging.error("Error during LLM call: %s", e)
            return ""
    raise Exception("Exceeded maximum retry attempts for LLM call")

def extract_text_from_file(file_path: str) -> str:
    text = ""
    fname = file_path.lower()
    try:
        if fname.endswith('.pdf'):
            text = extract_text_from_pdf(file_path)
        elif fname.endswith(('.xlsx', '.xls')):
            text = extract_text_from_excel(file_path)
        elif fname.endswith('.docx'):
            text = extract_text_from_docx(file_path)
        elif fname.endswith('.csv'):
            text = extract_text_from_csv(file_path)
        elif fname.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
            text = extract_text_from_image(file_path)
        else:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
    except Exception as e:
        return f"Error extracting text from {file_path}: {e}"

    # Add a final check to prevent empty strings
    if not text.strip():
        return f"Warning: No usable text extracted from {file_path}"
    return text

def extract_text_from_pdf(file_path: str) -> str:
    text = ""
    with open(file_path, 'rb') as file_content:
        reader = PdfReader(file_content)
        extracted = ""
        for page in reader.pages:
            pg_text = page.extract_text()
            if pg_text:
                extracted += pg_text + "\n"
    if len(extracted.strip()) < 50 and convert_from_path:
        pages = convert_from_path(file_path, dpi=200)
        if pytesseract:
            for p in pages:
                extracted += pytesseract.image_to_string(p) + "\n"
    return extracted

def extract_text_from_excel(file_path: str) -> str:
    df = pd.read_excel(file_path)
    numeric_columns = df.select_dtypes(include=['number']).columns

    # Convert numeric columns to numeric, handling errors
    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Use tab-separated format for consistent parsing
    return df.to_csv(sep='\t', index=False)

def extract_text_from_docx(file_path: str) -> str:
    document = docx.Document(file_path)
    return "\n".join(p.text for p in document.paragraphs)

def extract_text_from_csv(file_path: str) -> str:
    with open(file_path, 'r', encoding='utf-8') as file_content:
        reader = csv.reader(file_content)
        rows = []
        for row in reader:
            rows.append("\t".join(row))
        return "\n".join(rows)

def extract_text_from_image(file_path: str) -> str:
    if pytesseract:
        image = Image.open(file_path)
        return pytesseract.image_to_string(image)
    else:
        return "[No OCR support installed]"

def interpret_user_query(query: str) -> Dict[str, Any]:
    prompt = (
        f"Interpret the following user query and convert it into a valid Gmail search query by understanding the user's intent. "
        f"Use one of the following formats as appropriate:\n"
        f"   - from:example@gmail.com has:attachment\n"
        f"   - to:recipient@gmail.com after:2023/01/01 before:2023/12/31\n"
        f"   - subject:(\"meeting notes\" OR \"project update\")\n"
        f"   - is:unread label:work\n"
        f"   - \"important information\" -subject:\"important information\"\n"
        f"   - is:starred subject:report\n"
        f"   - in:spam from:spammer@example.com\n"
        f"   - has:attachment filename:pdf\n"
        f"   - \"follow up\" before:2023/01/01\n\n"
        f"Also, decide if the fetch_mode should be 'summary', 'full', or 'analyze', and set has_attachments to true if attachments are expected.\n\n"
        f"User Query: \"{query}\"\n"
        f"Return only the JSON with keys 'gmail_query', 'fetch_mode', and 'has_attachments'."
    )
    result = llm_predict_with_backoff(prompt)
    try:
        result = result.replace("`json", "").replace("`", "").strip()
        data = json.loads(result)
        if not any(op in data.get("gmail_query", "").lower() for op in ["from:", "to:", "after:", "before:", "subject:", "is:", "label:", "has:"]) and "\"" not in data.get("gmail_query", ""):
            if data.get("gmail_query", "").strip():
                data["gmail_query"] = f"\"{data['gmail_query']}\""
            else:
                raise ValueError("Interpreted query does not appear valid.")
        return data
    except Exception as e:
        logging.error("Error interpreting query: %s", e)
        return {
            "gmail_query": f"\"{query}\"",
            "fetch_mode": "summary",
            "has_attachments": "attachment" in query.lower()
        }

def get_gmail_credentials():
    creds = None
    token_path = 'token.json'
    if os.path.exists(token_path):
        with open(token_path, 'r') as token:
            creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, 'w') as token:
            token.write(creds.to_json())
    return creds

def fetch_gmail_data(query: Optional[str] = None, batch_size: int = 200) -> List[Dict[str, Any]]:
    """Fetch Gmail messages based on a query using OAuth credentials."""
    try:
        creds = None
        token_path = 'token.pickle'
        if os.path.exists(token_path):
            with open(token_path, 'rb') as token:
                creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_path, 'wb') as token:
                pickle.dump(creds, token)

        service = build('gmail', 'v1', credentials=creds)
        fetched_messages = []
        page_token = None
        while len(fetched_messages) < batch_size:
            if query:
                results = service.users().messages().list(
                    userId='me',
                    q=query,
                    maxResults=min(batch_size, 100),
                    pageToken=page_token
                ).execute()
            else:
                results = service.users().messages().list(
                    userId='me',
                    labelIds=['INBOX'],
                    maxResults=min(batch_size, 100),
                    pageToken=page_token
                ).execute()
            messages = results.get('messages', [])
            page_token = results.get('nextPageToken')
            for message in messages:
                msg = service.users().messages().get(
                    userId='me',
                    id=message['id'],
                    format='full'
                ).execute()
                fetched_messages.append(msg)
                if len(fetched_messages) >= batch_size:
                    break
            if not page_token:
                break
            time.sleep(1)
        return fetched_messages
    except Exception as e:
        logging.error("Error fetching Gmail data: %s", e)
        return []

def edit_file_content(initial_content: str, history: str = "") -> str:
    edit_win = tk.Toplevel()
    edit_win.title("Edit File Content / Analyze Data")
    edit_win.geometry("600x400")

    tk.Label(edit_win, text="Current Content (read-only):").pack(anchor="w", padx=10, pady=(10, 0))
    content_text = ScrolledText(edit_win, wrap=tk.WORD, height=10)
    content_text.pack(fill="both", expand=True, padx=10, pady=5)
    content_text.insert("1.0", initial_content)
    content_text.configure(state=tk.DISABLED)

    tk.Label(edit_win, text="Enter your instruction:").pack(anchor="w", padx=10, pady=(10, 0))
    instruction_entry = tk.Text(edit_win, wrap=tk.WORD, height=4)
    instruction_entry.pack(fill="x", padx=10, pady=5)

    result = {"edited": initial_content}

    def submit_edit():
        instruction = instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            messagebox.showerror("Error", "Instruction cannot be empty.")
            return
        prompt_context = f"Conversation history:\n{history}\n\n" if history else ""
        prompt = (
            prompt_context +
            "You are an advanced text editor, data analyst, and visualization assistant. "
            "If the instruction requires analysis or visualization, output the charts/graphs directly. "
            "Do not return raw Python code for visualization; instead, display the actual chart.\n\n"
            f"Instruction: {instruction}\n\n"
            f"Content:\n{initial_content}\n\n"
            "Return the final text, analysis, and if applicable, display the chart."
        )
        try:
            response = unified_llm.call(prompt)
            final_response = process_visualization_response(response)
            if not final_response.strip():
                messagebox.showinfo("LLM Output", "LLM returned no output.")
            result["edited"] = final_response.strip()
            edit_win.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to edit/analyze content: {e}")

    submit_btn = ttk.Button(edit_win, text="Submit Instruction", command=submit_edit)
    submit_btn.pack(padx=10, pady=10)

    edit_win.wait_window()
    return result["edited"]

def custom_chunk_text(text: str, method: str) -> List[str]:
    if method == "word":
        words = text.split()
        return [" ".join(words[i:i+50]) for i in range(0, len(words), 50)]
    elif method == "sentence":
        sentences = re.split(r'(?<=[.!?])\s+', text)
        return sentences
    elif method == "paragraph":
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        return paragraphs
    elif method == "pages":
        pages = re.split(r'\n{2,}', text)
        return pages
    elif method == "batches":
        pages = re.split(r'\n{2,}', text)
        batch_size = 3
        return ["\n\n".join(pages[i:i+batch_size]) for i in range(0, len(pages), batch_size)]
    elif method == "headings":
        lines = text.splitlines()
        chunks = []
        current_chunk = []
        for line in lines:
            if line.isupper() and len(line.strip()) > 3:
                if current_chunk:
                    chunks.append("\n".join(current_chunk))
                    current_chunk = []
                current_chunk.append(line)
            else:
                current_chunk.append(line)
        if current_chunk:
            chunks.append("\n".join(current_chunk))
        return chunks
    elif method == "headings_subheadings":
        lines = text.splitlines()
        chunks = []
        current_chunk = []
        for line in lines:
            if (line.isupper() and len(line.strip()) > 3) or (line.istitle() and len(line.split()) < 5):
                if current_chunk:
                    chunks.append("\n".join(current_chunk))
                    current_chunk = []
                current_chunk.append(line)
            else:
                current_chunk.append(line)
        if current_chunk:
            chunks.append("\n".join(current_chunk))
        return chunks
    else:
        return [text]

###############################################################################
# ScrollableFrame
###############################################################################
class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

###############################################################################
# DocumentStore
###############################################################################
class DocumentStore:
    def __init__(self, use_embeddings: bool = True, chunk_size: int = 500):
        self.use_embeddings = use_embeddings and (EMBEDDING_MODEL is not None)
        self.chunk_size = chunk_size
        self.docs: List[Tuple[str, str, str]] = []  # (short_name, chunk, original_name)
        self.embeddings = []
        self.short_to_original: Dict[str, str] = {}
        self.original_to_short: Dict[str, str] = {}
        self.file_counter = 1

    def _generate_short_name(self, original_name: str) -> str:
        short_name = f"File_{self.file_counter}"
        self.file_counter += 1
        self.short_to_original[short_name] = original_name
        self.original_to_short[original_name] = short_name
        return short_name

    def add_document_text(self, text: str, original_name: str):
        logging.info(f"Adding file to store: {original_name}")  # Debug line
        short_name = self._generate_short_name(original_name)
        chunks = self._chunk_text(text, self.chunk_size)
        for chunk in chunks:
            self.docs.append((short_name, chunk, original_name))
            if self.use_embeddings:
                emb = EMBEDDING_MODEL.encode(chunk, convert_to_tensor=True)
                self.embeddings.append(emb)
        logging.debug(f"Added document '{original_name}' with {len(chunks)} chunks to the document store.")

    def _chunk_text(self, text: str, chunk_size: int) -> List[str]:
        chunks = []
        start = 0
        while start < len(text):
            end = min(start + chunk_size, len(text))
            chunks.append(text[start:end])
            start = end
        return chunks

    def retrieve_relevant_chunks(self, query: str, top_k: int = 3) -> List[Tuple[str, str, str]]:
        if fuzz is not None:
            best_score = 0
            best_original = None
            for orig in self.original_to_short:
                score = fuzz.partial_ratio(orig.lower(), query.lower())
                if score > best_score:
                    best_score = score
                    best_original = orig
            if best_score > 70:
                short_name = self.original_to_short[best_original]
                return self._retrieve_from_short_name(short_name, query, top_k)
        else:
            for orig in self.original_to_short:
                if orig.lower() in query.lower():
                    short_name = self.original_to_short[orig]
                    return self._retrieve_from_short_name(short_name, query, top_k)
        if self.use_embeddings:
            query_emb = EMBEDDING_MODEL.encode(query, convert_to_tensor=True)
            logging.debug(f"Query Embedding Shape: {query_emb.shape}")

            if self.embeddings:
                embeddings_tensor = torch.stack(self.embeddings)
                logging.debug(f"Embeddings Tensor Shape: {embeddings_tensor.shape}")
            else:
                logging.warning("No embeddings found in the document store.")
                return []

            logging.debug(f"Calculating cosine similarity with query embedding shape: {query_emb.shape} and embeddings tensor shape: {embeddings_tensor.shape}")

            scores = util.pytorch_cos_sim(query_emb, embeddings_tensor)
            if scores.dim() > 1:
                logging.debug(f"Scores tensor shape before indexing: {scores.shape}")
                scores = scores[0]
                logging.debug(f"Scores tensor shape after indexing: {scores.shape}")

            logging.debug(f"Scores Shape: {scores.shape}")
            logging.debug(f"Scores Values: {scores}")
            top_results = scores.topk(k=min(top_k, len(scores))).indices.tolist()
            return [self.docs[idx] for idx in top_results]
        else:
            tokens_query = set(query.lower().split())
            def score(item: Tuple[str, str, str]) -> int:
                return len(tokens_query.intersection(set(item[1].lower().split())))
            ranked = sorted(self.docs, key=score, reverse=True)
            return [doc for doc in ranked[:top_k]]

    def _retrieve_from_short_name(self, short_name: str, query: str, top_k: int) -> List[Tuple[str, str, str]]:
        relevant_docs = [(sn, chunk, oname) for (sn, chunk, oname) in self.docs if sn == short_name]
        if not relevant_docs:
            return [("", "", f"No chunks found for file {short_name}")]

        if self.use_embeddings:
            relevant_indices = [i for i, (sn, _, _) in enumerate(self.docs) if sn == short_name]
            relevant_chunks = [self.docs[i][1] for i in relevant_indices]
            relevant_original_names = [self.docs[i][2] for i in relevant_indices]
            relevant_embs = [self.embeddings[i] for i in relevant_indices]

            # Ensure embeddings are stacked into a tensor
            if relevant_embs:
                embeddings_tensor = torch.stack(relevant_embs)
                query_emb = EMBEDDING_MODEL.encode(query, convert_to_tensor=True)
                scores = util.pytorch_cos_sim(query_emb, embeddings_tensor)[0]
                top_results = scores.topk(k=min(top_k, len(scores))).indices.tolist()
                return [(short_name, relevant_chunks[idx], relevant_original_names[idx]) for idx in top_results]

        return relevant_docs[:top_k]

    def get_all_files(self) -> List[str]:
        """Return all unique filenames stored."""
        return list(set(original_name for _, _, original_name in self.docs))

###############################################################################
# RAGChatBot
###############################################################################
class RAGChatBot:
    def __init__(self, doc_store: DocumentStore):
        self.doc_store = doc_store
        self.conversation_history = ""
        self.retrieved_context_content = ""

    def add_document(self, text: str, original_name: str):
        self.doc_store.add_document_text(text, original_name)
        logging.debug(f"Added document: {original_name} to the document store.")

    def chat(self, user_query: str, top_k: int = 3) -> str:
        if "list all the files present in the context" in user_query.lower():
            # Directly fetch all stored filenames
            all_files = self.doc_store.get_all_files()
            context_block = "\n".join([f"* {fname}" for fname in all_files]) if all_files else "No files found."
        else:
            relevant_chunks_with_names = self.doc_store.retrieve_relevant_chunks(user_query, top_k=top_k)
            context_block = "\n".join([f"* {original_name}" for _, _, original_name in relevant_chunks_with_names])

        self.retrieved_context_content = context_block
        prompt = (
            "You are a helpful assistant with access to local documents via the doc store. "
            "If the user references a file by name, use that file's data.\n\n"
            f"The files present in the context are:\n{context_block}\n\n"
            f"User: {user_query}\nAssistant:"
        )
        response = llm_predict_with_backoff(prompt)
        self.conversation_history += f"\nUser: {user_query}\nAssistant: {response}"
        return process_visualization_response(response)

    def get_last_retrieved_context(self) -> str:
        return self.retrieved_context_content

    def format_response(self, context_block: str) -> str:
        return f"The following files are present in the context:\n{context_block}"

###############################################################################
# Visualization and Code Execution
###############################################################################
def verify_and_run_code(code: str, namespace: Dict[str, Any]) -> str:
    try:
        output_buffer = io.StringIO()
        error_buffer = io.StringIO()
        namespace['plt'] = plt

        def run_code():
            with contextlib.redirect_stdout(output_buffer), contextlib.redirect_stderr(error_buffer):
                exec(code, namespace)
            fig = plt.gcf()
            root.after(0, lambda: _display_chart(fig, output_buffer.getvalue(), error_buffer.getvalue()))

        threading.Thread(target=run_code).start()
        return "[Visualization Generated]\nWaiting for chart..."

    except Exception as e:
        error_text = error_buffer.getvalue()
        import traceback
        traceback_str = traceback.format_exc()
        error_message = f"[Error executing code: {e}]\n[Traceback]:\n{traceback_str}\n[Standard Error output]:\n{error_text}"
        logging.error(f"Error in verify_and_run_code: {error_message}")
        return error_message

def _display_chart(fig, output_text, error_text):
    def show_chart():
        chart_win = tk.Toplevel()
        chart_win.title("Generated Visualization")
        canvas = FigureCanvasTkAgg(fig, master=chart_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        full_output = "[Visualization Generated]\n" + output_text + ("\n[Errors]:\n" + error_text if error_text else "")
        set_code_output(full_output)

    root.after(0, show_chart)


def process_visualization_response(response: str) -> str:
    """
    This updated function removes any hard-coded chart logic or default column checks.
    It simply looks for code blocks (triple backticks) in the LLM response and returns them
    so the AI Code Editor can do the logical/syntax checks and generate the correct visualization.
    """
    import re
    logging.debug(f"Processing visualization response: {response}")

    # Extract code blocks enclosed in triple backticks.
    code_blocks = re.findall(r"```(.*?)```", response, re.DOTALL)
    if code_blocks:
        # Trim whitespace from each code block
        trimmed_blocks = [block.strip() for block in code_blocks]
        logging.debug(f"Code blocks found: {trimmed_blocks}")
        # Return only the extracted code blocks so the AI Code Editor can process them
        return "\n".join(trimmed_blocks)

    # If no code blocks are found, return the original response as-is.
    return response

###############################################################################
# Custom Editors
###############################################################################
class CustomTextEditor(tk.Toplevel):
    def __init__(self, parent, file_path, master_app):
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"Text Editor - {os.path.basename(file_path)}")
        self.geometry("1000x800")

        self.file_path = file_path
        self.current_content = extract_text_from_file(file_path)

        self.paned_window = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        self.text_area = ScrolledText(left_frame, wrap=tk.WORD, height=25)
        self.text_area.pack(expand=True, fill='both', padx=10, pady=(10,0))
        self.text_area.insert('1.0', self.current_content)

        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)
        self.save_button = ttk.Button(nl_frame, text="Save Changes", command=self.save_changes_to_file)
        self.save_button.pack(pady=5)

        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED)
        self.rag_output_display.pack(expand=True, fill='both')

        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent")
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED)
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGTextEditorAgent(self)

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return
        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("")
        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            response = self.rag_agent.run_instruction(instruction, self.current_content)
            self.set_rag_output(response)
            if "[EDITED_CONTENT]" in response:
                edited_content = response.split("[EDITED_CONTENT]")[1].split("[/EDITED_CONTENT]")[0].strip()
                self.update_text_content(edited_content)
            context_content = self.master_app.chatbot.get_last_retrieved_context()
            self.set_context_output(context_content)
        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")

    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL)
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED)

    def set_context_output(self, text):
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)

    def update_text_content(self, new_content):
        self.text_area.delete("1.0", tk.END)
        self.text_area.insert("1.0", new_content)
        self.current_content = new_content

    def save_changes_to_file(self):
        new_content = self.text_area.get("1.0", tk.END)
        try:
            with open(self.file_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            self.current_content = new_content
            messagebox.showinfo("File Saved", f"Changes saved to: {os.path.basename(self.file_path)}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Error saving file: {e}")

class CustomImageViewer(tk.Toplevel):
    def __init__(self, parent, file_path, master_app):
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"Image Viewer - {os.path.basename(file_path)}")
        self.geometry("900x700")

        self.file_path = file_path
        self.ocr_text = extract_text_from_image(file_path)  # Extract OCR text upfront

        self.paned_window = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        try:
            img = Image.open(file_path)
            img.thumbnail((800, 600))  # Resize image to fit within the window
            img_tk = ImageTk.PhotoImage(img)
            self.image_label = tk.Label(left_frame, image=img_tk)
            self.image_label.image = img_tk
            self.image_label.pack(padx=10, pady=(10, 0))
        except Exception as e:
            tk.Label(left_frame, text=f"Error opening image: {e}").pack(padx=10, pady=10)

        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        # Add Natural Language Query Section at the top of the right frame
        nl_frame = ttk.LabelFrame(right_frame, text="OCR & Analysis Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)

        # RAG Agent Output
        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED)
        self.rag_output_display.pack(expand=True, fill='both')

        # Context Accessed by Agent
        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent")
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED)
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGFileViewerAgent(self)

        # Add OCR Text Preview
        self.ocr_preview_frame = ttk.LabelFrame(left_frame, text="OCR Text Preview")
        self.ocr_preview_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.ocr_preview = ScrolledText(self.ocr_preview_frame, wrap=tk.WORD, height=10, state=tk.DISABLED)
        self.ocr_preview.pack(expand=True, fill='both')
        self.update_ocr_preview()

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return
        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("")
        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            # Use the pre-extracted OCR text
            response = self.rag_agent.run_instruction(instruction, self.ocr_text)
            self.set_rag_output(response)
            # Show the OCR text as context
            self.set_context_output(self.ocr_text)
        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")

    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL)
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED)

    def set_context_output(self, text):
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)

    def update_ocr_preview(self):
        self.ocr_preview.config(state=tk.NORMAL)
        self.ocr_preview.delete("1.0", tk.END)
        self.ocr_preview.insert("1.0", self.ocr_text)
        self.ocr_preview.config(state=tk.DISABLED)

class CustomExcelViewer(tk.Toplevel):
    def __init__(self, parent, file_path, master_app):
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"Excel Viewer - {os.path.basename(file_path)}")
        self.geometry("900x700")

        self.file_path = file_path

        self.paned_window = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        self.tree = ttk.Treeview(left_frame)
        self.tree.pack(expand=True, fill='both', padx=10, pady=(10,0))
        try:
            self.wb = openpyxl.load_workbook(file_path)
            self.sheet = self.wb.active
            headers = [cell.value for cell in self.sheet[1]]
            self.tree["columns"] = headers
            self.tree["show"] = "headings"
            for col in headers:
                self.tree.heading(col, text=col)
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", tk.END, values=row)
        except Exception as e:
            tk.Label(left_frame, text=f"Error opening Excel file: {e}").pack(padx=10, pady=10)

        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)
        self.save_button = ttk.Button(nl_frame, text="Save Changes", command=self.save_changes_to_file)
        self.save_button.pack(pady=5)

        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED)
        self.rag_output_display.pack(expand=True, fill='both')

        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent")
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED)
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGFileViewerAgent(self)

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return
        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("")
        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            file_content = extract_text_from_file(self.file_path)
            response = self.rag_agent.run_instruction(instruction, file_content)
            self.set_rag_output(response)
            if "[EDITED_CONTENT]" in response:
                edited_content = response.split("[EDITED_CONTENT]")[1].split("[/EDITED_CONTENT]")[0].strip()
                self.update_excel_content(edited_content)
            context_content = self.master_app.chatbot.get_last_retrieved_context()
            self.set_context_output(context_content)
        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")

    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL)
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED)

    def set_context_output(self, text):
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)

    def update_excel_content(self, new_content):
        lines = new_content.split('\n')
        if not lines:
            return

        headers = lines[0].split('\t')  # Split headers by tab
        header_count = len(headers)
        formatted_data = []
        for line in lines[1:]:  # Skip header line
            if not line.strip():
                continue
            row = line.split('\t')  # Split each row by tab
            if len(row) != header_count:
                # Handle mismatched columns (e.g., log error or skip)
                logging.error(f"Mismatched columns: Expected {header_count}, got {len(row)}")
                continue

            # Safely format the last column (assuming it's the profit)
            try:
                profit = float(row[-1])
                row[-1] = f"{profit:.2f}"
            except (ValueError, IndexError):
                row[-1] = "0.00"  # Default to 0.00 if conversion fails

            formatted_data.append(row)

        # Update the Treeview with the new formatted data
        for row in formatted_data:
            self.tree.insert("", tk.END, values=row)

        # Update the Excel file
        for i, row in enumerate(formatted_data):
            for j, value in enumerate(row):
                self.sheet.cell(row=i+2, column=j+1, value=value)

        messagebox.showinfo("File Saved", f"Changes saved to: {os.path.basename(self.file_path)}")

    def save_changes_to_file(self):
        try:
            self.wb.save(self.file_path)
            messagebox.showinfo("File Saved", f"Changes saved to: {os.path.basename(self.file_path)}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Error saving file: {e}")

class CustomPDFViewer(tk.Toplevel):
    def __init__(self, parent, file_path, master_app):
        super().__init__(parent)
        self.master_app = master_app
        self.title(f"PDF Viewer - {os.path.basename(file_path)}")
        self.geometry("900x700")

        self.file_path = file_path
        self.image_based_pdf = False

        self.paned_window = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill="both", expand=True)

        left_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(left_frame, weight=2)

        self.text_area = ScrolledText(left_frame, wrap=tk.WORD, height=25)
        self.text_area.pack(expand=True, fill='both', padx=10, pady=(10,0))
        content = extract_text_from_file(file_path)
        self.text_area.insert('1.0', content)
        self.text_area.config(state=tk.DISABLED)

        nl_frame = ttk.LabelFrame(left_frame, text="Natural Language Instructions")
        nl_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(nl_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_instruction_entry = tk.Text(nl_frame, height=3, wrap=tk.WORD)
        self.nl_instruction_entry.pack(fill="x")
        self.nl_button = ttk.Button(nl_frame, text="Submit Instruction", command=self.process_nl_instruction)
        self.nl_button.pack(pady=5)

        right_frame = ttk.Frame(self.paned_window)
        self.paned_window.add(right_frame, weight=1)

        self.rag_output_frame = ttk.LabelFrame(right_frame, text="RAG Agent Output")
        self.rag_output_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.rag_output_display = ScrolledText(self.rag_output_frame, height=15, wrap=tk.WORD, state=tk.DISABLED)
        self.rag_output_display.pack(expand=True, fill='both')

        self.context_frame = ttk.LabelFrame(right_frame, text="Context Accessed by Agent")
        self.context_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_display = ScrolledText(self.context_frame, height=10, wrap=tk.WORD, state=tk.DISABLED)
        self.context_display.pack(expand=True, fill='both')

        self.rag_agent = RAGFileViewerAgent(self)

    def process_nl_instruction(self):
        instruction = self.nl_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            return
        self.nl_instruction_entry.delete("1.0", tk.END)
        self.set_rag_output("Processing instruction...")
        self.set_context_output("")
        threading.Thread(target=self._run_rag_instruction, args=(instruction,)).start()

    def _run_rag_instruction(self, instruction):
        try:
            file_content = extract_text_from_file(self.file_path)
            response = self.rag_agent.run_instruction(instruction, file_content)
            self.set_rag_output(response)
            context_content = self.master_app.chatbot.get_last_retrieved_context()
            self.set_context_output(context_content)
        except Exception as e:
            self.set_rag_output(f"Error processing instruction: {e}")
            self.set_context_output(f"Error retrieving context: {e}")

    def set_rag_output(self, text):
        self.rag_output_display.config(state=tk.NORMAL)
        self.rag_output_display.delete("1.0", tk.END)
        self.rag_output_display.insert("1.0", text)
        self.rag_output_display.config(state=tk.DISABLED)

    def set_context_output(self, text):
        self.context_display.config(state=tk.NORMAL)
        self.context_display.delete("1.0", tk.END)
        self.context_display.insert("1.0", text)
        self.context_display.config(state=tk.DISABLED)

def open_custom_editor(parent, file_path, master_app):
    if file_path.lower().endswith(('.txt', '.csv', '.docx')):
        CustomTextEditor(parent, file_path, master_app)
    elif file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
        CustomImageViewer(parent, file_path, master_app)
    elif file_path.lower().endswith(('.xlsx', '.xls')):
        CustomExcelViewer(parent, file_path, master_app)
    elif file_path.lower().endswith('.pdf'):
        CustomPDFViewer(parent, file_path, master_app)
    else:
        messagebox.showinfo("Info", "No custom viewer available for this file type. Opening as text editor.")
        CustomTextEditor(parent, file_path, master_app)

###############################################################################
# RAG Agents
###############################################################################
class RAGTextEditorAgent:
    def __init__(self, editor: CustomTextEditor):
        self.editor = editor

    def run_instruction(self, instruction: str, file_content: str) -> str:
        prompt = (
            "You are an expert text editor AI. Analyze the content and follow the user's instruction.\n"
            "You have access to relevant context from emails and dropbox files if applicable. Use this context to better understand the user's request and file content.\n"
            "If the instruction is to edit the content, perform the edit and clearly indicate the edited content in your response like this: '[EDITED_CONTENT]\n[the edited content here]\n[/EDITED_CONTENT]'. "
            "If the instruction is for analysis, perform the analysis and present the result clearly.\n"
            f"Instruction: {instruction}\n"
            f"File Content:\n{file_content}\n"
            "Response:"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw)
        return response_processed

class RAGFileViewerAgent:
    def __init__(self, viewer):
        self.viewer = viewer

    def run_instruction(self, instruction: str, file_content: str) -> str:
        prompt = (
            "You are an expert AI assistant for file analysis. You are viewing a file and will respond to the user's questions or "
            "perform analysis based on the file content. You have access to relevant context from emails and other files in the "
            "File Box to enrich your understanding.\n\n"

            "If the user asks you to modify the file (e.g., reformat columns, round values, etc.), you must provide a preview of "
            "the updated content **without** complaining about missing columns. Use the exact changes the user requests.\n\n"

            "Ensure numerical columns remain numeric. Avoid inserting strings into numeric columns like 'Profit'.\n\n"
            f"Instruction: {instruction}\n\n"
            f"File Content:\n{file_content}\n\n"
            "Response:"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw)
        return response_processed

class RAGCodeEditorAgent:
    def __init__(self, editor: 'AICodeEditor'):
        self.editor = editor
        self.last_retrieved_context = ""  # Store the last retrieved context

    def run_instruction(self, instruction: str, code_content: str) -> str:
        doc_store = self.editor.master_app.chatbot.doc_store
        relevant_chunks_with_names = doc_store.retrieve_relevant_chunks(instruction, top_k=3)
        context_block = "\n\n".join([f"Relevant chunk from '{original_name}':\n{chunk}" for _, chunk, original_name in relevant_chunks_with_names])
        self.last_retrieved_context = context_block  # Store for display

        prompt = (
            "You are an expert AI code editor and data visualization assistant. "
            "You can generate, analyze, debug, and run Python code. "
            "You can also create data visualizations using matplotlib.pyplot. \n"
            "You have access to relevant information from files in the system (see context below), emails, and dropbox files. "
            "Use this context to better understand the user's request and the code. \n"
            f"Context:\n{context_block}\n\n"
            f"Current Code in Editor:\n{code_content}\n\n"
            f"Instruction: {instruction}\n"
            "Response:\n"
        )

        # Log the instruction and code content
        logging.debug(f"Running instruction: {instruction}")
        logging.debug(f"Code content:\n{code_content}")

        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw)
        return response_processed

class RAGFileBoxAgent:
    def __init__(self, chatbot: RAGChatBot):
        self.chatbot = chatbot

    def run_instruction(self, instruction: str) -> str:
        relevant_chunks_with_names = self.chatbot.doc_store.retrieve_relevant_chunks(instruction, top_k=5)
        context_block = "\n\n".join([f"Relevant chunk from '{original_name}':\n{chunk}" for _, chunk, original_name in relevant_chunks_with_names])

        prompt = (
            "You are a highly capable and versatile AI assistant expert in analyzing and manipulating various file types within a File Box environment. "
            "You have access to a wide range of tools and functionalities, including:\n"
            "   - Analyzing text, images, spreadsheets, and PDF documents.\n"
            "   - Summarizing information, extracting key insights, and answering questions based on file content.\n"
            "   - Generating code for data visualization and analysis using Python and matplotlib (indicate code blocks with triple backticks).\n"
            "   - Reasoning and problem-solving using the information available in the files.\n"
            "   - Accessing relevant context from emails and other files in the File Box to enrich your understanding and responses.\n"
            "   - When generating multiple charts, create a new figure for each plot using plt.figure(). Do not reuse the same figure object.\n"
            "You should leverage these capabilities to understand user instructions related to files in File Box and provide comprehensive and helpful responses.\n"
            "If the user asks to edit a text-based file, provide instructions on how to open it in the text editor and use natural language instructions within the editor for modification.\n"
            "Do not access external websites or local file system other than provided file context.\n\n"
            f"Context from File Box and Emails:\n{context_block}\n\n"
            f"User Instruction: {instruction}\n"
            "Response:"
        )
        response_raw = llm_predict_with_backoff(prompt)
        response_processed = process_visualization_response(response_raw)
        self.chatbot.retrieved_context_content = context_block
        return response_processed

###############################################################################
# AI Code Editor
###############################################################################
class AICodeEditor(ttk.Frame):
    def __init__(self, parent, master_app):
        super().__init__(parent)
        self.master_app = master_app
        self.parent = parent
        self.rag_code_agent = RAGCodeEditorAgent(self)
        self.code_namespace = {}
        self.file_context_content = ""

        self.main_paned_window = ttk.PanedWindow(self, orient=tk.VERTICAL)
        self.main_paned_window.pack(fill="both", expand=True)

        self.top_paned_window = ttk.PanedWindow(self.main_paned_window, orient=tk.HORIZONTAL)
        self.main_paned_window.add(self.top_paned_window, weight=1)

        self.input_frame = ttk.Frame(self.top_paned_window, relief=tk.SOLID, borderwidth=1)
        self.top_paned_window.add(self.input_frame, weight=2)
        ttk.Label(self.input_frame, text="Code Input (Python):", font=("TkDefaultFont", 9, 'bold')).pack(anchor="w", padx=5, pady=(5,0))
        self.code_input_area = ScrolledText(self.input_frame, wrap=tk.WORD, height=15, borderwidth=0)
        self.code_input_area.pack(expand=True, fill='both', padx=5, pady=5)

        self.output_frame = ttk.Frame(self.top_paned_window, relief=tk.SOLID, borderwidth=1)
        self.top_paned_window.add(self.output_frame, weight=2)
        ttk.Label(self.output_frame, text="Code Output/Visualization:", font=("TkDefaultFont", 9, 'bold')).pack(anchor="w", padx=5, pady=(5,0))
        self.code_output_display = ScrolledText(self.output_frame, wrap=tk.WORD, height=10, state=tk.DISABLED, borderwidth=0)
        self.code_output_display.pack(expand=True, fill='both', padx=5, pady=5)

        self.bottom_frame = ttk.Frame(self.main_paned_window)
        self.main_paned_window.add(self.bottom_frame, weight=1)

        nl_code_frame = ttk.LabelFrame(self.bottom_frame, text="Natural Language Instructions for Code Editor")
        nl_code_frame.pack(fill="both", expand=True, padx=10, pady=5)
        tk.Label(nl_code_frame, text="Enter instruction:").pack(anchor="w")
        self.nl_code_instruction_entry = tk.Text(nl_code_frame, height=3, wrap=tk.WORD)
        self.nl_code_instruction_entry.pack(fill="x")
        self.nl_code_button = ttk.Button(nl_code_frame, text="Submit Code Instruction", command=self.process_code_nl_instruction)
        self.nl_code_button.pack(pady=5)

        self.context_code_frame = ttk.LabelFrame(self.bottom_frame, text="Context Accessed by Agent")
        self.context_code_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.context_code_display = ScrolledText(self.context_code_frame, height=8, wrap=tk.WORD, state=tk.DISABLED)
        self.context_code_display.pack(expand=True, fill='both')

        button_input_frame = ttk.Frame(self.input_frame)
        button_input_frame.pack(fill="x", pady=5, padx=5)

        self.run_button = ttk.Button(button_input_frame, text="Run Code", command=self.run_code)
        self.run_button.pack(side="left", padx=5)
        self.clear_code_button = ttk.Button(button_input_frame, text="Clear Code", command=self.clear_code)
        self.clear_code_button.pack(side="left", padx=5)
        self.save_visual_button = ttk.Button(button_input_frame, text="Save Visualization", command=self.save_visualization, state=tk.DISABLED)
        self.save_visual_button.pack(side="left", padx=5)
        self.load_file_context_button = ttk.Button(button_input_frame, text="Load File Context", command=self.load_file_context)
        self.load_file_context_button.pack(side="left", padx=5)

    def load_file_context(self):
        dropbox_files = [self.master_app.filebox_file_listbox.get(idx) for idx in self.master_app.filebox_file_listbox.curselection()]
        gmail_messages = self.master_app.gmail_messages

        context_text = ""
        for filepath in dropbox_files:
            context_text += f"\n--- Content from File Box file: {os.path.basename(filepath)} ---\n"
            context_text += extract_text_from_file(filepath)

        if gmail_messages:
            for msg in gmail_messages:
                if 'payload' in msg and 'parts' in msg['payload']:
                    for part in msg['payload']['parts']:
                        if 'filename' in part and part['filename']:
                            file_name = part['filename']
                            file_data_b64 = part['body']['data'] if 'body' in part and 'data' in part['body'] else None
                            if file_data_b64:
                                try:
                                    file_data = base64.urlsafe_b64decode(file_data_b64)
                                    temp_dir = tempfile.mkdtemp()
                                    file_path = os.path.join(temp_dir, file_name)
                                    with open(file_path, 'wb') as temp_file:
                                        temp_file.write(file_data)
                                    context_text += f"\n--- Content from Gmail attachment: {file_name} ---\n"
                                    context_text += extract_text_from_file(file_path)
                                except Exception as e:
                                    logging.error(f"Error processing Gmail attachment {file_name} for code editor context: {e}")

        self.file_context_content = context_text
        messagebox.showinfo("Context Loaded", "File context loaded for AI Code Editor. The AI agent can now access content from selected File Box and Gmail files when you give instructions.")

    def process_code_nl_instruction(self):
        instruction = self.nl_code_instruction_entry.get("1.0", tk.END).strip()
        if not instruction:
            logging.warning("No instruction provided for code processing.")
            return
        self.nl_code_instruction_entry.delete("1.0", tk.END)
        self.set_code_output("Processing code instruction...")
        self.set_context_code_output("")
        logging.info(f"Processing code instruction: {instruction}")
        threading.Thread(target=self._run_code_nl_instruction, args=(instruction,)).start()

    def _run_code_nl_instruction(self, instruction):
        code_content = self.code_input_area.get("1.0", tk.END)
        logging.debug(f"Code content before running instruction: {code_content}")
        try:
            response = self.rag_code_agent.run_instruction(instruction, code_content)
            logging.debug(f"Response from code agent: {response}")
            self.set_code_output(response)
            if "[Visualization generated]" in response:
                self.enable_save_visual_button()
            else:
                self.disable_save_visual_button()
            self.set_context_code_output(self.rag_code_agent.last_retrieved_context)
        except Exception as e:
            logging.error(f"Error processing code instruction: {e}")
            self.set_code_output(f"Error processing code instruction: {e}")
            self.set_context_code_output(f"Error retrieving context: {e}")
            self.disable_save_visual_button()

    def run_code(self):
        code_to_run = self.code_input_area.get("1.0", tk.END)
        self.set_code_output("Running code...")
        self.disable_save_visual_button()
        logging.info("Running code...")
        threading.Thread(target=self._execute_code, args=(code_to_run,)).start()

    def _execute_code(self, code):
        try:
            output = verify_and_run_code(code, self.code_namespace)
            self.set_code_output(output)
            logging.info("Code executed successfully.")
            if "[Visualization Generated]" in output:
                self.enable_save_visual_button()
            else:
                self.disable_save_visual_button()
            self.set_context_code_output("No context retrieved for direct code execution.")
        except Exception as e:
            logging.error(f"Error executing code: {e}")
            self.set_code_output(f"Error executing code: {e}")
            self.set_context_code_output(f"Error retrieving context: {e}")
            self.disable_save_visual_button()

    def set_code_output(self, text):
        self.code_output_display.config(state=tk.NORMAL)
        self.code_output_display.delete("1.0", tk.END)
        self.code_output_display.insert("1.0", text)
        self.code_output_display.config(state=tk.DISABLED)

    def set_context_code_output(self, text):
        self.context_code_display.config(state=tk.NORMAL)
        self.context_code_display.delete("1.0", tk.END)
        self.context_code_display.insert("1.0", text)
        self.context_code_display.config(state=tk.DISABLED)

    def enable_save_visual_button(self):
        self.save_visual_button.config(state=tk.NORMAL)

    def disable_save_visual_button(self):
        self.save_visual_button.config(state=tk.DISABLED)

    def save_visualization(self):
        if self.code_namespace.get('fig'):
            file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG files", "*.png"), ("All files", "*.*")])
            if file_path:
                self.code_namespace['fig'].savefig(file_path)
                messagebox.showinfo("Save Visualization", "Visualization saved successfully.")
        else:
            messagebox.showerror("Error", "No visualization to save.")

    def clear_code(self):
        self.code_input_area.delete("1.0", tk.END)
        self.set_code_output("")
        self.set_context_code_output("")
        self.disable_save_visual_button()

###############################################################################
# Main Application
###############################################################################
class MultiDocRAGApp:
    def __init__(self, parent: ttk.Frame, root: tk.Tk, document_store: DocumentStore):
        self.parent = parent
        self.root = root
        self.root.title("Multi-Document RAG Application")
        self.root.geometry("1200x800")

        self.chatbot = RAGChatBot(document_store)  # Use the shared DocumentStore
        self.filebox_rag_agent = RAGFileBoxAgent(self.chatbot)
        self.filebox_files_dir = "FileBox_files"
        os.makedirs(self.filebox_files_dir, exist_ok=True)
        self.gmail_messages = []
        self.last_code_block = None

        self.notebook = ttk.Notebook(self.parent)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self.chat_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.chat_tab, text="RAG Chatbot")
        self.create_chat_tab_ui(self.chat_tab)

        self.filebox_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.filebox_tab, text="File Box")
        self.create_filebox_tab_ui(self.filebox_tab)

        self.code_editor_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.code_editor_tab, text="AI Code Editor")
        self.create_code_editor_tab_ui(self.code_editor_tab)

        self.context_display_filebox = ScrolledText(self.filebox_tab, wrap=tk.WORD, height=10, state=tk.DISABLED)
        self.context_display_filebox.pack(fill=tk.BOTH, expand=True)

    def create_chat_tab_ui(self, parent_tab):
        file_frame = ttk.Frame(parent_tab)
        file_frame.pack(fill=tk.X)
        self.file_listbox = tk.Listbox(file_frame, height=5)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Button(file_frame, text="Load Files", command=self.load_local_files).pack(side=tk.LEFT)

        self.chat_display_area = ScrolledText(parent_tab, wrap=tk.WORD, height=20, state=tk.DISABLED)
        self.chat_display_area.pack(fill=tk.BOTH, expand=True)
        self.chat_input_entry = ttk.Entry(parent_tab)
        self.chat_input_entry.pack(fill=tk.X)
        self.chat_input_entry.bind("<Return>", self.send_chat_message)

    def load_local_files(self):
        files = filedialog.askopenfilenames()
        for file_path in files:
            file_name = os.path.basename(file_path)
            self.file_listbox.insert(tk.END, file_name)
            text = extract_text_from_file(file_path)
            self.chatbot.add_document(text, file_name)

    def send_chat_message(self, event=None):
        message = self.chat_input_entry.get()
        if message:
            self.chat_display_area.config(state=tk.NORMAL)
            self.chat_display_area.insert(tk.END, f"User: {message}\n")
            self.chat_display_area.insert(tk.END, "Assistant: Thinking...\n")
            self.chat_display_area.config(state=tk.DISABLED)
            self.chat_input_entry.delete(0, tk.END)
            threading.Thread(target=self._get_chat_response, args=(message,)).start()

    def _get_chat_response(self, message):
        response = self.chatbot.chat(message)
        self.chat_display_area.config(state=tk.NORMAL)
        self.chat_display_area.delete("end-2l", "end-1l")
        self.chat_display_area.insert(tk.END, f"Assistant: {response}\n")
        self.chat_display_area.config(state=tk.DISABLED)

    def create_filebox_tab_ui(self, parent_tab):
        filebox_frame = ttk.Frame(parent_tab)
        filebox_frame.pack(fill=tk.BOTH, expand=True)

        left_frame = ttk.Frame(filebox_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.filebox_file_listbox = tk.Listbox(left_frame, selectmode=tk.MULTIPLE)
        self.filebox_file_listbox.pack(fill=tk.BOTH, expand=True)
        ttk.Button(left_frame, text="Upload Files", command=self.upload_filebox_files).pack()
        ttk.Button(left_frame, text="Open Selected", command=self.open_selected_files).pack()

        right_frame = ttk.Frame(filebox_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        self.filebox_chat_area = ScrolledText(right_frame, wrap=tk.WORD, height=20, state=tk.DISABLED)
        self.filebox_chat_area.pack(fill=tk.BOTH, expand=True)
        self.filebox_instruction_area = tk.Text(right_frame, height=3)
        self.filebox_instruction_area.pack(fill=tk.X)
        button_frame = ttk.Frame(right_frame)
        button_frame.pack(fill=tk.X)
        ttk.Button(button_frame, text="Submit", command=self.process_filebox_instruction).pack(side=tk.LEFT)
        self.send_to_editor_button = ttk.Button(button_frame, text="Send Code to Editor", command=self.send_code_to_editor, state=tk.DISABLED)
        self.send_to_editor_button.pack(side=tk.LEFT)

    def upload_filebox_files(self):
        files = filedialog.askopenfilenames()
        for file_path in files:
            try:
                dest_path = os.path.join(self.filebox_files_dir, os.path.basename(file_path))
                shutil.copy(file_path, dest_path)
                self.filebox_file_listbox.insert(tk.END, dest_path)

                # Extract text and add to document store UNCONDITIONALLY
                text = extract_text_from_file(dest_path)
                self.chatbot.add_document(text, os.path.basename(file_path))
                logging.info(f"Added document to store: {os.path.basename(file_path)}")
            except Exception as e:
                logging.error(f"Failed to process {file_path}: {e}")

    def open_selected_files(self):
        selected = self.filebox_file_listbox.curselection()
        for idx in selected:
            file_path = self.filebox_file_listbox.get(idx)
            open_custom_editor(self.parent, file_path, self)

    def process_filebox_instruction(self):
        instruction = self.filebox_instruction_area.get("1.0", tk.END).strip()
        if instruction:
            self.filebox_chat_area.config(state=tk.NORMAL)
            self.filebox_chat_area.insert(tk.END, f"User: {instruction}\n")
            self.filebox_chat_area.insert(tk.END, "Assistant: Thinking...\n", "thinking")
            self.filebox_chat_area.config(state=tk.DISABLED)
            self.filebox_instruction_area.delete("1.0", tk.END)

            def get_response():
                try:
                    response = self.filebox_rag_agent.run_instruction(instruction)
                    logging.debug(f"FileBox response: {response}")  # Add logging for debugging
                    self.root.after(0, self.update_filebox_display, "Assistant:", response)
                    context_content = self.chatbot.retrieved_context_content
                    self.root.after(0, self.update_context_display_filebox, context_content)
                except Exception as e:
                    self.root.after(0, self.update_filebox_display, "Assistant:", f"Error: {e}")
                    self.root.after(0, self.update_context_display_filebox, f"Error retrieving context: {e}")

            threading.Thread(target=get_response).start()

    def update_filebox_display(self, sender, message):
        self.filebox_chat_area.after(0, self.filebox_chat_area.config, {'state': tk.NORMAL})
        # Remove "Thinking..." if present
        thinking_ranges = self.filebox_chat_area.tag_ranges("thinking")
        if thinking_ranges:
            start, end = thinking_ranges[0], thinking_ranges[1]
            self.filebox_chat_area.after(0, self.filebox_chat_area.delete, start, end)
        self.filebox_chat_area.after(0, self.filebox_chat_area.insert, tk.END, f"{sender} {message}\n")
        self.filebox_chat_area.after(0, self.filebox_chat_area.config, {'state': tk.DISABLED})
        self.filebox_chat_area.after(0, self.filebox_chat_area.see, tk.END)  # Scroll to the bottom

    def update_context_display_filebox(self, context):
        self.context_display_filebox.after(0, self.context_display_filebox.config, {'state': tk.NORMAL})
        self.context_display_filebox.after(0, self.context_display_filebox.delete, "1.0", tk.END)
        self.context_display_filebox.after(0, self.context_display_filebox.insert, "1.0", context)
        self.context_display_filebox.after(0, self.context_display_filebox.config, {'state': tk.DISABLED})

    def send_code_to_editor(self):
        if self.last_code_block:
            self.code_editor.code_input_area.delete("1.0", tk.END)
            self.code_editor.code_input_area.insert("1.0", self.last_code_block)
            self.notebook.select(self.code_editor_tab)

    def create_gmail_tab_ui(self, parent_tab):
        search_frame = ttk.Frame(parent_tab)
        search_frame.pack(fill=tk.X)
        self.gmail_search_entry = ttk.Entry(search_frame)
        self.gmail_search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(search_frame, text="Search Gmail", command=self.search_gmail).pack(side=tk.LEFT)

        self.gmail_listbox = tk.Listbox(parent_tab, height=10)
        self.gmail_listbox.pack(fill=tk.BOTH, expand=True)
        self.gmail_listbox.bind('<<ListboxSelect>>', self.display_gmail_content)

        self.gmail_content_area = ScrolledText(parent_tab, wrap=tk.WORD, height=10, state=tk.DISABLED)
        self.gmail_content_area.pack(fill=tk.BOTH, expand=True)

    def search_gmail(self):
        query = self.gmail_search_entry.get()
        self.gmail_messages = fetch_gmail_data(query)
        self.gmail_listbox.delete(0, tk.END)
        for msg in self.gmail_messages:
            subject = next((h['value'] for h in msg['payload']['headers'] if h['name'] == 'Subject'), 'No Subject')
            self.gmail_listbox.insert(tk.END, subject)

    def display_gmail_content(self, event):
        selection = self.gmail_listbox.curselection()
        if selection:
            msg = self.gmail_messages[selection[0]]
            content = ""
            if 'parts' in msg['payload']:
                for part in msg['payload']['parts']:
                    if part['mimeType'] == 'text/plain' and 'data' in part['body']:
                        content += base64.urlsafe_b64decode(part['body']['data']).decode('utf-8')
            self.gmail_content_area.config(state=tk.NORMAL)
            self.gmail_content_area.delete("1.0", tk.END)
            self.gmail_content_area.insert("1.0", content)
            self.gmail_content_area.config(state=tk.DISABLED)

    def create_code_editor_tab_ui(self, parent_tab):
        self.code_editor = AICodeEditor(parent_tab, self)
        self.code_editor.pack(fill=tk.BOTH, expand=True)

class GmailDataFetcherTool(BaseTool):
    name: str = "gmail_data_fetcher"
    description: str = (
        "Fetches Gmail data and attachments, processes emails using an LLM, and allows viewing a summary with "
        "buttons to view full email content and open attachments in a custom editor."
    )
    config: Dict[str, Any] = Field(default_factory=dict)
    conversation_history: List[str] = Field(default_factory=list)
    cache: ClassVar[Dict[str, Any]] = Field(default_factory=dict)

    def __init__(self, **data: Any):
        super().__init__(**data)
        object.__setattr__(self, "config", self.load_external_config("search_config.json"))

    def load_external_config(self, filename: str) -> Dict[str, Any]:
        if os.path.exists(filename):
            with open(filename, "r") as file:
                return json.load(file)
        return {}

    def parse_chunking_method(self, query: str) -> Optional[str]:
        q = query.lower()
        if "chunk by word" in q:
            return "word"
        elif "chunk by sentence" in q:
            return "sentence"
        elif "chunk by paragraph" in q:
            return "paragraph"
        elif "chunk by pages" in q:
            return "pages"
        elif "chunk by batches" in q:
            return "batches"
        elif "chunk by headings_subheadings" in q:
            return "headings_subheadings"
        elif "chunk by headings" in q:
            return "headings"
        else:
            return None

    def extract_email_text(self, msg: Dict[str, Any], chunk_method: Optional[str] = None,
                           for_summary: bool = True) -> str:
        try:
            email_text = ""
            payload = msg.get('payload', {})
            if 'parts' in payload:
                for part in payload['parts']:
                    if part.get('mimeType', '') == 'text/plain':
                        data = part['body'].get('data')
                        if data:
                            email_text += base64.urlsafe_b64decode(data.encode('UTF-8')).decode('UTF-8')
            else:
                if payload.get('mimeType') == 'text/plain':
                    data = payload['body'].get('data')
                    if data:
                        email_text = base64.urlsafe_b64decode(data.encode('UTF-8')).decode('UTF-8')
            return email_text
        except Exception as e:
            logging.error("Error extracting email text: %s", e)
            return f"Error extracting email text: {e}"

    def get_attachment_info(self, msg: Dict[str, Any]) -> List[Tuple[str, str, Dict[str, Any]]]:
        attachments = []
        payload = msg.get('payload', {})
        if 'parts' in payload:
            for part in payload['parts']:
                if part.get('filename'):
                    attachments.append((part.get('filename'), msg['id'], part))
        return attachments

    def download_attachment(self, message_id: str, part: Dict[str, Any], download_path: str):
        try:
            service = build('gmail', 'v1', credentials=self.get_gmail_credentials())
            attachment_id = part['body']['attachmentId']
            response = service.users().messages().attachments().get(
                userId='me',
                messageId=message_id,
                id=attachment_id
            ).execute()
            file_data = base64.urlsafe_b64decode(response['data'].encode('UTF-8'))
            file_path = os.path.join(download_path, part['filename'])
            with open(file_path, 'wb') as file:
                file.write(file_data)
            logging.info("Attachment downloaded to %s", file_path)
        except Exception as e:
            logging.error("Error downloading attachment: %s", e)
            return f"Error downloading attachment: {e}"

    def get_gmail_credentials(self):
        creds = None
        token_path = 'token.pickle'
        if os.path.exists(token_path):
            with open(token_path, 'rb') as token:
                creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_path, 'wb') as token:
                pickle.dump(creds, token)
        return creds

    def extract_text_from_file(self, file_path: str, file_name: str) -> str:
        return extract_text_from_file(file_path)

    def _run(self, query: str) -> List[Dict[str, Any]]:
        try:
            logging.info(f"Running GmailDataFetcherTool with query: {query}")
            intent = interpret_user_query(query)
            logging.info(f"Interpreted intent: {intent}")
            gmail_query = intent.get("gmail_query", query)
            chunk_method = self.parse_chunking_method(query)
            logging.info(f"Chunk method: {chunk_method}")
            messages = fetch_gmail_data(query=gmail_query, batch_size=10)
            logging.info(f"Fetched {len(messages)} messages.")
            email_info_list = []
            for msg in messages:
                full_text = self.extract_email_text(msg, chunk_method, for_summary=True)
                summary_prompt = f"Summarize the following email content in one or two sentences:\n{full_text}"
                summary = llm_predict_with_backoff(summary_prompt)
                attachments = self.get_attachment_info(msg)
                email_info_list.append({
                    "summary": summary,
                    "full": full_text,
                    "attachments": attachments
                })
            return email_info_list
        except Exception as e:
            logging.error(f"Tool execution failed: {e}", exc_info=True)
            return [{"summary": f"Error processing Gmail data: {e}", "full": "", "attachments": []}]

def open_custom_editor_attachment(attachment_info: Tuple[str, str, Dict[str, Any]]) -> None:
    filename, message_id, part = attachment_info
    temp_dir = tempfile.mkdtemp()
    GmailDataFetcherTool().download_attachment(message_id, part, temp_dir)
    file_path = os.path.join(temp_dir, filename)
    open_custom_editor(file_path)

class GmailDataFetcherUI:
    def __init__(self, parent: tk.Widget, doc_app=None, chatbot=None, document_store=None):
        self.parent = parent
        self.doc_app = doc_app  # Reference to MultiDocRAGApp instance
        self.chatbot = chatbot  # Reference to chatbot instance
        self.fetcher_tool = GmailDataFetcherTool()
        self.document_store = document_store  # Reference to shared DocumentStore

        # Input Frame
        self.input_frame = ttk.Frame(parent)
        self.input_frame.pack(fill="x", padx=10, pady=5)
        self.query_label = ttk.Label(self.input_frame, text="Enter Gmail Query:")
        self.query_label.pack(side="left")
        self.query_entry = ttk.Entry(self.input_frame, width=50)
        self.query_entry.pack(side="left", padx=5)
        self.fetch_button = ttk.Button(self.input_frame, text="Fetch and Process", command=self.fetch_and_process)
        self.fetch_button.pack(side="left", padx=5)
        self.advanced_analysis_var = tk.BooleanVar(value=False)
        self.advanced_check = ttk.Checkbutton(self.input_frame, text="Enable Advanced Analysis", variable=self.advanced_analysis_var)
        self.advanced_check.pack(side="left", padx=5)

        # Emails Panel
        self.emails_frame = ttk.Frame(parent)
        self.emails_frame.pack(fill="both", expand=True)
        # Status Bar
        self.status_bar = ttk.Label(parent, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side="bottom", fill="x")

    def fetch_and_process(self):
        query = self.query_entry.get()
        if not query:
            messagebox.showerror("Error", "Query cannot be empty.")
            return
        self.status_bar.config(text="Fetching and processing...")
        self.parent.update_idletasks()
        try:
            intent = interpret_user_query(query)
            logging.info(f"Interpreted intent: {intent}")
            gmail_query = intent.get("gmail_query", query)
            email_info_list = self.fetcher_tool._run(gmail_query)
            for widget in self.emails_frame.winfo_children():
                widget.destroy()
            for email_info in email_info_list:
                frame = ttk.Frame(self.emails_frame, relief=tk.RIDGE, borderwidth=1)
                frame.pack(fill="x", pady=5)
                summary_label = ttk.Label(frame, text="Summary: " + email_info["summary"], wraplength=600, justify=tk.LEFT)
                summary_label.pack(side="top", anchor="w", padx=5, pady=2)
                btn_full = ttk.Button(frame, text="View Full Email", command=lambda text=email_info["full"]: self.open_full_email(text))
                btn_full.pack(side="top", anchor="w", padx=5, pady=2)
                if email_info["attachments"]:
                    attach_label = ttk.Label(frame, text="Attachments:")
                    attach_label.pack(side="top", anchor="w", padx=5)
                    for attachment_info in email_info["attachments"]:
                        filename, message_id, part = attachment_info
                        btn_attach = ttk.Button(
                            frame,
                            text=f"Open {filename}",
                            command=lambda ai=(filename, message_id, part): self.open_attachment(ai)
                        )
                        btn_attach.pack(side="top", anchor="w", padx=15, pady=2)
            self.status_bar.config(text="Done.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.status_bar.config(text=f"Error: {e}")
        finally:
            self.parent.update_idletasks()

    def open_full_email(self, full_text: str):
        top = tk.Toplevel(self.parent)
        top.title("Full Email Content")
        text_widget = ScrolledText(top, wrap=tk.WORD, height=25, width=80)
        text_widget.pack(fill="both", expand=True, padx=10, pady=10)
        text_widget.insert("1.0", full_text)

    def open_attachment(self, attachment_info: Tuple[str, str, Dict[str, Any]]):
        try:
            filename, message_id, part = attachment_info
            temp_dir = tempfile.mkdtemp()
            self.fetcher_tool.download_attachment(message_id, part, temp_dir)
            file_path = os.path.join(temp_dir, filename)

            # 1. Extract text from the downloaded file
            file_text = extract_text_from_file(file_path)
            logging.debug(f"Extracted text from {filename}: {file_text[:100]}...")  # Log the first 100 characters

            # 2. Add that text to the doc store (via the MultiDocRAGApp instance)
            self.doc_app.chatbot.add_document(file_text, filename)
            logging.debug(f"Added document to store: {filename}")

            # 3. Open the file in a custom editor
            open_custom_editor(self.parent, file_path, self.doc_app)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open attachment: {e}")
            logging.error(f"Error opening attachment: {e}")

class MainApp(ttk.Frame):
    def __init__(self, root: tk.Tk):
        super().__init__(root)
        root.title("Unified Application: Gmail & Document Analysis")
        root.geometry("1000x800")

        # Initialize a shared DocumentStore
        self.document_store = DocumentStore(use_embeddings=True)

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True)

        # Initialize Document Chat Tab first
        self.doc_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.doc_frame, text="Document Chat")
        self.doc_ui = MultiDocRAGApp(self.doc_frame, root, self.document_store)  # Pass the shared DocumentStore

        # Now initialize Gmail Tab
        self.gmail_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.gmail_frame, text="Gmail Data")
        self.gmail_ui = GmailDataFetcherUI(self.gmail_frame, doc_app=self.doc_ui, chatbot=self.doc_ui.chatbot, document_store=self.document_store)  # Pass the shared DocumentStore

def preprocess_data_before_analysis(df: pd.DataFrame, columns_to_check=None) -> pd.DataFrame:
    """
    Checks for non-numeric values in the specified columns (or all columns if none specified).
    Prompts the user how to handle them before proceeding with analysis.
    """

    if columns_to_check is None:
        # Default: check all columns in df
        columns_to_check = df.columns

    # Create a hidden root window for dialogs
    root = tk.Tk()
    root.withdraw()

    for col in columns_to_check:
        # Convert everything we can to numeric; 'errors="coerce"' turns invalid entries into NaN
        coerced = pd.to_numeric(df[col], errors='coerce')
        invalid_mask = coerced.isna() & df[col].notna()  # True where original was non-numeric

        if invalid_mask.any():
            # We found some non-numeric values
            unique_invalid = df.loc[invalid_mask, col].unique()
            # Prompt the user for how to handle these values
            message = (
                f"Column '{col}' contains non-numeric values: {list(unique_invalid)}\n\n"
                "How would you like to handle them?\n\n"
                "1) Replace with 0\n"
                "2) Replace with mean of valid values\n"
                "3) Drop rows containing non-numeric values\n"
                "4) Cancel (no changes)\n"
            )
            choice = simpledialog.askstring("Data Preprocessing", message)

            if choice == "1":
                # Replace invalid with 0
                df[col] = coerced.fillna(0)
                messagebox.showinfo("Data Preprocessing", f"All non-numeric values in '{col}' replaced with 0.")
            elif choice == "2":
                # Replace invalid with mean of valid values
                valid_mean = coerced.dropna().mean()
                df[col] = coerced.fillna(valid_mean)
                messagebox.showinfo("Data Preprocessing", f"All non-numeric values in '{col}' replaced with mean={valid_mean:.2f}.")
            elif choice == "3":
                # Drop rows containing invalid values
                df.drop(df[invalid_mask].index, inplace=True)
                messagebox.showinfo("Data Preprocessing", f"Dropped rows with non-numeric values in '{col}'.")
            else:
                # Cancel, do nothing
                messagebox.showinfo("Data Preprocessing", "Operation cancelled. No changes applied.")
                return df  # Return immediately if user cancels

    return df

# Example usage in your analysis function
def analyze_data():
    df = pd.read_excel("my_spreadsheet.xlsx")
    df = preprocess_data_before_analysis(df)  # This will prompt the user if any non-numeric data is found
    # Now proceed with analysis, plotting, etc.

if __name__ == "__main__":
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()
